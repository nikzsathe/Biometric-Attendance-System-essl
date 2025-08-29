from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session
from flask_mail import Mail, Message
import sqlite3
from datetime import datetime, timedelta
import os
from zk import ZK
import socket
import hashlib
import secrets
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

# Initialize Flask app
app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # Change this in production

# Email configuration
app.config['MAIL_SERVER'] = 'smtp.gmail.com'  # Change to your email server
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'your-email@gmail.com'  # Change to your email
app.config['MAIL_PASSWORD'] = 'your-app-password'  # Change to your app password
app.config['MAIL_DEFAULT_SENDER'] = 'your-email@gmail.com'  # Change to your email

# Initialize Flask-Mail
mail = Mail(app)

# Initialize scheduler for automated emails
scheduler = BackgroundScheduler()
scheduler.start()

# Company name constant
COMPANY_NAME = "Absolute Global Outsourcing"

# Context processor to make COMPANY_NAME available in all templates
@app.context_processor
def inject_company_name():
    return {'COMPANY_NAME': COMPANY_NAME}

# Database path
DB_PATH = 'attendance.db'

# Device configuration
DEFAULT_DEVICE_IP = '192.168.1.201'
DEFAULT_DEVICE_PORT = 4370

# Admin credentials (you can change these)
ADMIN_USERNAME = 'admin'
ADMIN_PASSWORD = 'admin123'  # Change this to a secure password

# Developer name
DEVELOPER_NAME = 'Nikhil Sathe'

def hash_password(password):
    """Hash a password using SHA-256"""
    return hashlib.sha256(password.encode()).hexdigest()

def verify_admin(username, password):
    """Verify admin credentials"""
    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        return True
    return False

def login_required(f):
    """Decorator to require admin login for protected routes"""
    def decorated_function(*args, **kwargs):
        if 'admin_logged_in' not in session:
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

def get_db_connection():
    """Create a database connection"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def send_email_notification(subject, recipients, html_content, attachments=None):
    """Send email notification with optional attachments"""
    try:
        msg = Message(
            subject=subject,
            recipients=recipients,
            html=html_content
        )
        
        if attachments:
            for attachment in attachments:
                msg.attach(attachment['filename'], attachment['content_type'], attachment['data'])
        
        mail.send(msg)
        return True
    except Exception as e:
        print(f"Error sending email: {e}")
        return False

def generate_attendance_summary_email():
    """Generate daily attendance summary email"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get today's attendance summary
        today = datetime.now().strftime('%Y-%m-%d')
        cursor.execute('''
            SELECT u.name, u.company_name, a.status, a.check_in, a.check_out, a.working_hours
            FROM users u
            LEFT JOIN attendance a ON u.userid = a.userid AND DATE(a.timestamp) = ?
            ORDER BY u.company_name, u.name
        ''', (today,))
        
        records = cursor.fetchall()
        conn.close()
        
        if not records:
            return None
        
        # Generate HTML content
        html_content = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                .header {{ background-color: #366092; color: white; padding: 20px; text-align: center; }}
                .summary {{ margin: 20px 0; }}
                table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                .present {{ color: green; }}
                .absent {{ color: red; }}
                .late {{ color: orange; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Daily Attendance Summary</h1>
                <h2>{COMPANY_NAME}</h2>
                <p>Date: {today}</p>
            </div>
            
            <div class="summary">
                <h3>Attendance Overview</h3>
                <table>
                    <tr>
                        <th>Employee</th>
                        <th>Company</th>
                        <th>Status</th>
                        <th>Check In</th>
                        <th>Check Out</th>
                        <th>Working Hours</th>
                    </tr>
        """
        
        for record in records:
            status_class = 'present' if record['status'] == 'present' else 'absent'
            html_content += f"""
                    <tr>
                        <td>{record['name']}</td>
                        <td>{record['company_name']}</td>
                        <td class="{status_class}">{record['status'] or 'Not marked'}</td>
                        <td>{record['check_in'] or '-'}</td>
                        <td>{record['check_out'] or '-'}</td>
                        <td>{record['working_hours'] or 0} hours</td>
                    </tr>
            """
        
        html_content += """
                </table>
            </div>
            
            <div style="margin-top: 30px; padding: 20px; background-color: #f9f9f9;">
                <p><strong>Note:</strong> This is an automated report generated by the Attendance System.</p>
                <p>Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
        </body>
        </html>
        """.format(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        return html_content
        
    except Exception as e:
        print(f"Error generating attendance summary: {e}")
        return None

def generate_salary_summary_email(month, year):
    """Generate monthly salary summary email"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get salary data for the month
        cursor.execute('''
            SELECT u.name, u.company_name, u.monthly_salary,
                   COUNT(DISTINCT DATE(a.timestamp)) as present_days,
                   SUM(a.working_hours) as total_hours,
                   SUM(CASE WHEN a.working_hours > 8 THEN a.working_hours - 8 ELSE 0 END) as overtime_hours
            FROM users u
            LEFT JOIN attendance a ON u.userid = a.userid 
                AND strftime('%Y-%m', a.timestamp) = ?
            GROUP BY u.userid, u.name, u.company_name, u.monthly_salary
            ORDER BY u.company_name, u.name
        ''', (f"{year}-{month:02d}",))
        
        records = cursor.fetchall()
        conn.close()
        
        if not records:
            return None
        
        # Calculate totals
        total_employees = len(records)
        total_salary = sum(record['monthly_salary'] for record in records)
        total_overtime = sum(record['overtime_hours'] for record in records)
        
        # Generate HTML content
        html_content = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                .header {{ background-color: #366092; color: white; padding: 20px; text-align: center; }}
                .summary {{ margin: 20px 0; }}
                .totals {{ background-color: #f0f8ff; padding: 20px; margin: 20px 0; }}
                table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Monthly Salary Summary</h1>
                <h2>{COMPANY_NAME}</h2>
                <p>Period: {month} {year}</p>
            </div>
            
            <div class="totals">
                <h3>Summary</h3>
                <p><strong>Total Employees:</strong> {total_employees}</p>
                <p><strong>Total Monthly Salary:</strong> ₹{total_salary:,.2f}</p>
                <p><strong>Total Overtime Hours:</strong> {total_overtime:.2f} hours</p>
            </div>
            
            <div class="summary">
                <h3>Employee Details</h3>
                <table>
                    <tr>
                        <th>Employee</th>
                        <th>Company</th>
                        <th>Monthly Salary</th>
                        <th>Present Days</th>
                        <th>Total Hours</th>
                        <th>Overtime Hours</th>
                    </tr>
        """
        
        for record in records:
            html_content += f"""
                    <tr>
                        <td>{record['name']}</td>
                        <td>{record['company_name']}</td>
                        <td>₹{record['monthly_salary']:,.2f}</td>
                        <td>{record['present_days'] or 0}</td>
                        <td>{record['total_hours'] or 0:.2f}</td>
                        <td>{record['overtime_hours'] or 0:.2f}</td>
                    </tr>
            """
        
        html_content += """
                </table>
            </div>
            
            <div style="margin-top: 30px; padding: 20px; background-color: #f9f9f9;">
                <p><strong>Note:</strong> This is an automated salary report generated by the Attendance System.</p>
                <p>Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
        </body>
        </html>
        """.format(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        return html_content
        
    except Exception as e:
        print(f"Error generating salary summary: {e}")
        return None

def save_email_config(email, password):
    """Save email configuration to database"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if config exists
        cursor.execute('SELECT id FROM email_config LIMIT 1')
        existing = cursor.fetchone()
        
        if existing:
            # Update existing config
            cursor.execute('''UPDATE email_config 
                             SET email = ?, password = ?, updated_date = CURRENT_TIMESTAMP 
                             WHERE id = ?''', (email, password, existing['id']))
        else:
            # Insert new config
            cursor.execute('''INSERT INTO email_config (email, password) 
                             VALUES (?, ?)''', (email, password))
        
        conn.commit()
        conn.close()
        return True
        
    except Exception as e:
        print(f"Error saving email config: {e}")
        return False

def load_email_config():
    """Load email configuration from database"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT email, password FROM email_config LIMIT 1')
        config = cursor.fetchone()
        conn.close()
        
        if config:
            return {
                'email': config['email'],
                'password': config['password']
            }
        return None
        
    except Exception as e:
        print(f"Error loading email config: {e}")
        return None

# Analytics Functions
def get_analytics_quick_stats(company=None, from_date=None, to_date=None, shift=None):
    """Get quick statistics for analytics dashboard"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Date filter
        if from_date and to_date:
            start_date = from_date
            end_date = to_date
        else:
            # Default to last 30 days if no dates provided
            from datetime import datetime, timedelta
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
            start_date = start_date.strftime('%Y-%m-%d')
            end_date = end_date.strftime('%Y-%m-%d')
        
        # Company filter
        company_filter = ""
        if company:
            company_filter = f"AND u.company_name = '{company}'"
        
        # Shift filter
        shift_filter = ""
        if shift:
            shift_filter = f"AND u.shift_type = '{shift}'"
        
        # Total employees
        cursor.execute(f'''
            SELECT COUNT(*) as total
            FROM users u
            WHERE 1=1 {company_filter} {shift_filter}
        ''')
        total_employees = cursor.fetchone()['total']
        
        # Present today
        today = datetime.now().strftime('%Y-%m-%d')
        cursor.execute(f'''
            SELECT COUNT(DISTINCT u.userid) as present
            FROM users u
            JOIN attendance a ON u.userid = a.userid
            WHERE DATE(a.timestamp) = ? {company_filter} {shift_filter}
        ''', (today,))
        present_today = cursor.fetchone()['present']
        
        # Late today (assuming 9 AM start for day shift, 7 PM for night shift)
        cursor.execute(f'''
            SELECT COUNT(DISTINCT u.userid) as late
            FROM users u
            JOIN attendance a ON u.userid = a.userid
            WHERE DATE(a.timestamp) = ? {company_filter} {shift_filter}
            AND (
                (u.shift_type = 'day' AND TIME(a.timestamp) > '09:00:00') OR
                (u.shift_type = 'night' AND TIME(a.timestamp) > '19:00:00')
            )
        ''', (today,))
        late_today = cursor.fetchone()['late']
        
        # Average working hours
        cursor.execute(f'''
            SELECT AVG(a.working_hours) as avg_hours
            FROM users u
            JOIN attendance a ON u.userid = a.userid
            WHERE DATE(a.timestamp) >= ? AND DATE(a.timestamp) <= ? {company_filter} {shift_filter}
            AND a.working_hours > 0
        ''', (start_date, end_date))
        result = cursor.fetchone()
        avg_working_hours = result['avg_hours'] if result['avg_hours'] else 0
        
        conn.close()
        
        return {
            'totalEmployees': total_employees,
            'presentToday': present_today,
            'lateToday': late_today,
            'avgWorkingHours': round(avg_working_hours, 1)
        }
        
    except Exception as e:
        print(f"Error getting quick stats: {e}")
        return {
            'totalEmployees': 0,
            'presentToday': 0,
            'lateToday': 0,
            'avgWorkingHours': 0
        }

def get_attendance_trend_data(company=None, from_date=None, to_date=None, shift=None):
    """Get attendance trend data for charts"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Date filter
        if from_date and to_date:
            start_date = datetime.strptime(from_date, '%Y-%m-%d')
            end_date = datetime.strptime(to_date, '%Y-%m-%d')
        else:
            # Default to last 30 days if no dates provided
            from datetime import datetime, timedelta
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
        
        # Company and shift filters
        company_filter = f"AND u.company_name = '{company}'" if company else ""
        shift_filter = f"AND u.shift_type = '{shift}'" if shift else ""
        
        # Get daily attendance data
        cursor.execute(f'''
            SELECT 
                DATE(a.timestamp) as date,
                COUNT(DISTINCT u.userid) as present,
                (SELECT COUNT(*) FROM users u2 WHERE 1=1 {company_filter} {shift_filter}) - COUNT(DISTINCT u.userid) as absent
            FROM users u
            JOIN attendance a ON u.userid = a.userid
            WHERE DATE(a.timestamp) >= ? AND DATE(a.timestamp) <= ? {company_filter} {shift_filter}
            GROUP BY DATE(a.timestamp)
            ORDER BY date
        ''', (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
        
        records = cursor.fetchall()
        conn.close()
        
        # Prepare chart data
        labels = []
        present_data = []
        absent_data = []
        
        current_date = start_date
        while current_date <= end_date:
            date_str = current_date.strftime('%Y-%m-%d')
            labels.append(date_str)
            
            # Find data for this date
            record = next((r for r in records if r['date'] == date_str), None)
            if record:
                present_data.append(record['present'])
                absent_data.append(record['absent'])
            else:
                present_data.append(0)
                absent_data.append(0)
            
            current_date += timedelta(days=1)
        
        return {
            'labels': labels,
            'present': present_data,
            'absent': absent_data
        }
        
    except Exception as e:
        print(f"Error getting attendance trend: {e}")
        return {'labels': [], 'present': [], 'absent': []}

def get_company_distribution_data():
    """Get company distribution data for pie chart"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT company_name, COUNT(*) as count
            FROM users
            GROUP BY company_name
            ORDER BY count DESC
        ''')
        
        records = cursor.fetchall()
        conn.close()
        
        labels = [record['company_name'] for record in records]
        values = [record['count'] for record in records]
        
        return {
            'labels': labels,
            'values': values
        }
        
    except Exception as e:
        print(f"Error getting company distribution: {e}")
        return {'labels': [], 'values': []}

def get_working_hours_data(company=None, from_date=None, to_date=None, shift=None):
    """Get working hours data for bar chart"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Date filter
        if from_date and to_date:
            start_date = from_date
            end_date = to_date
        else:
            # Default to last 30 days if no dates provided
            from datetime import datetime, timedelta
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
            start_date = start_date.strftime('%Y-%m-%d')
            end_date = end_date.strftime('%Y-%m-%d')
        
        # Company and shift filters
        company_filter = f"AND u.company_name = '{company}'" if company else ""
        shift_filter = f"AND u.shift_type = '{shift}'" if shift else ""
        
        # Get weekly working hours average
        cursor.execute(f'''
            SELECT 
                strftime('%W', a.timestamp) as week,
                AVG(a.working_hours) as avg_hours
            FROM users u
            JOIN attendance a ON u.userid = a.userid
            WHERE DATE(a.timestamp) >= ? AND DATE(a.timestamp) <= ? {company_filter} {shift_filter}
            AND a.working_hours > 0
            GROUP BY strftime('%W', a.timestamp)
            ORDER BY week
            LIMIT 8
        ''', (start_date, end_date))
        
        records = cursor.fetchall()
        conn.close()
        
        labels = [f"Week {record['week']}" for record in records]
        values = [round(record['avg_hours'], 1) for record in records]
        
        return {
            'labels': labels,
            'values': values
        }
        
    except Exception as e:
        print(f"Error getting working hours data: {e}")
        return {'labels': [], 'values': []}

def get_overtime_trend_data(company=None, from_date=None, to_date=None, shift=None):
    """Get overtime trend data for area chart"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Date filter
        if from_date and to_date:
            start_date = datetime.strptime(from_date, '%Y-%m-%d')
            end_date = datetime.strptime(to_date, '%Y-%m-%d')
        else:
            # Default to last 30 days if no dates provided
            from datetime import datetime, timedelta
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
        
        # Company and shift filters
        company_filter = f"AND u.company_name = '{company}'" if company else ""
        shift_filter = f"AND u.shift_type = '{shift}'" if shift else ""
        
        # Get daily overtime data
        cursor.execute(f'''
            SELECT 
                DATE(a.timestamp) as date,
                SUM(CASE WHEN a.working_hours > 8 THEN a.working_hours - 8 ELSE 0 END) as overtime
            FROM users u
            JOIN attendance a ON u.userid = a.userid
            WHERE DATE(a.timestamp) >= ? AND DATE(a.timestamp) <= ? {company_filter} {shift_filter}
            GROUP BY DATE(a.timestamp)
            ORDER BY date
        ''', (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
        
        records = cursor.fetchall()
        conn.close()
        
        # Prepare chart data
        labels = []
        values = []
        
        current_date = start_date
        while current_date <= end_date:
            date_str = current_date.strftime('%Y-%m-%d')
            labels.append(date_str)
            
            # Find data for this date
            record = next((r for r in records if r['date'] == date_str), None)
            if record:
                values.append(round(record['overtime'], 1))
            else:
                values.append(0)
            
            current_date += timedelta(days=1)
        
        return {
            'labels': labels,
            'values': values
        }
        
    except Exception as e:
        print(f"Error getting overtime trend: {e}")
        return {'labels': [], 'values': []}

def get_performance_table_data(company=None, from_date=None, to_date=None, shift=None):
    """Get employee performance data for table"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Date filter
        if from_date and to_date:
            start_date = from_date
            end_date = to_date
            # Calculate total days between dates
            from datetime import datetime
            start_dt = datetime.strptime(from_date, '%Y-%m-%d')
            end_dt = datetime.strptime(to_date, '%Y-%m-%d')
            total_days = (end_dt - start_dt).days + 1
        else:
            # Default to last 30 days if no dates provided
            from datetime import datetime, timedelta
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
            start_date = start_date.strftime('%Y-%m-%d')
            end_date = end_date.strftime('%Y-%m-%d')
            total_days = 30
        
        # Company and shift filters
        company_filter = f"AND u.company_name = '{company}'" if company else ""
        shift_filter = f"AND u.shift_type = '{shift}'" if shift else ""
        
        # Get employee performance data
        cursor.execute(f'''
            SELECT 
                u.name,
                u.company_name,
                COUNT(DISTINCT DATE(a.timestamp)) as present_days,
                AVG(a.working_hours) as avg_working_hours,
                SUM(CASE WHEN a.working_hours > 8 THEN a.working_hours - 8 ELSE 0 END) as overtime_hours,
                COUNT(CASE WHEN 
                    (u.shift_type = 'day' AND TIME(a.timestamp) > '09:00:00') OR
                    (u.shift_type = 'night' AND TIME(a.timestamp) > '19:00:00')
                THEN 1 END) as late_arrivals
            FROM users u
            LEFT JOIN attendance a ON u.userid = a.userid 
                AND DATE(a.timestamp) >= ? AND DATE(a.timestamp) <= ? {company_filter} {shift_filter}
            WHERE 1=1 {company_filter} {shift_filter}
            GROUP BY u.userid, u.name, u.company_name
            ORDER BY u.company_name, u.name
        ''', (start_date, end_date))
        
        records = cursor.fetchall()
        conn.close()
        
        # Calculate performance metrics
        performance_data = []
        for record in records:
            # Calculate attendance rate
            attendance_rate = (record['present_days'] / total_days) * 100 if total_days > 0 else 0
            
            # Calculate performance score (0-10)
            attendance_score = min(attendance_rate / 10, 4)  # Max 4 points
            working_hours_score = min((record['avg_working_hours'] or 0) / 8, 3)  # Max 3 points
            overtime_score = min((record['overtime_hours'] or 0) / 10, 2)  # Max 2 points
            punctuality_score = max(0, 1 - ((record['late_arrivals'] or 0) / 10))  # Max 1 point
            
            performance_score = attendance_score + working_hours_score + overtime_score + punctuality_score
            
            performance_data.append({
                'name': record['name'],
                'company_name': record['company_name'],
                'attendanceRate': round(attendance_rate, 1),
                'avgWorkingHours': round(record['avg_working_hours'] or 0, 1),
                'overtimeHours': round(record['overtime_hours'] or 0, 1),
                'lateArrivals': record['late_arrivals'] or 0,
                'performanceScore': round(performance_score, 1)
            })
        
        return performance_data
        
    except Exception as e:
        print(f"Error getting performance data: {e}")
        return []

def setup_db():
    """Initialize database and create tables if they don't exist"""
    conn = get_db_connection()
    c = conn.cursor()
    
    # Create users table
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        userid INTEGER PRIMARY KEY,
        name TEXT NOT NULL,
        company_name TEXT DEFAULT 'Absolute Global Outsourcing',
        shift_start_time TEXT DEFAULT '09:00',
        shift_end_time TEXT DEFAULT '18:00',
        shift_type TEXT DEFAULT 'day',
        working_hours_per_day REAL DEFAULT 8.0,
        monthly_salary REAL DEFAULT 15000.0,
        created_date TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # Create attendance table
    c.execute('''CREATE TABLE IF NOT EXISTS attendance (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        userid INTEGER NOT NULL,
        timestamp TEXT NOT NULL,
        check_in TEXT,
        check_out TEXT,
        working_hours REAL DEFAULT 0.0,
        status TEXT DEFAULT 'present',
        FOREIGN KEY (userid) REFERENCES users (userid)
    )''')
    
    # Create companies table
    c.execute('''CREATE TABLE IF NOT EXISTS companies (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        description TEXT,
        created_date TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # Create holidays table
    c.execute('''CREATE TABLE IF NOT EXISTS holidays (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT UNIQUE NOT NULL,
        name TEXT NOT NULL,
        description TEXT,
        is_public_holiday BOOLEAN DEFAULT 1,
        created_date TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # Create email_config table
    c.execute('''CREATE TABLE IF NOT EXISTS email_config (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        email TEXT NOT NULL,
        password TEXT NOT NULL,
        smtp_server TEXT DEFAULT 'smtp.gmail.com',
        smtp_port INTEGER DEFAULT 587,
        smtp_security TEXT DEFAULT 'TLS',
        updated_date TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # Create attendance_marking table
    c.execute('''CREATE TABLE IF NOT EXISTS attendance_marking (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        userid INTEGER NOT NULL,
        date TEXT NOT NULL,
        status TEXT NOT NULL,
        working_hours REAL DEFAULT 0.0,
        overtime_hours REAL DEFAULT 0.0,
        late_minutes INTEGER DEFAULT 0,
        remarks TEXT,
        marked_by TEXT DEFAULT 'system',
        created_date TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (userid) REFERENCES users (userid),
        UNIQUE(userid, date)
    )''')
    
    # Create salary_calculations table
    c.execute('''CREATE TABLE IF NOT EXISTS salary_calculations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        userid INTEGER NOT NULL,
        month TEXT NOT NULL,
        year INTEGER NOT NULL,
        total_days INTEGER DEFAULT 0,
        present_days INTEGER DEFAULT 0,
        absent_days INTEGER DEFAULT 0,
        leave_days INTEGER DEFAULT 0,
        total_working_hours REAL DEFAULT 0.0,
        overtime_hours REAL DEFAULT 0.0,
        basic_salary REAL DEFAULT 0.0,
        overtime_pay REAL DEFAULT 0.0,
        deductions REAL DEFAULT 0.0,
        net_salary REAL DEFAULT 0.0,
        calculated_date TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (userid) REFERENCES users (userid),
        UNIQUE(userid, month, year)
    )''')
    
    # Check if working_hours column exists in attendance table
    c.execute("PRAGMA table_info(attendance)")
    columns = [column[1] for column in c.fetchall()]
    
    if 'working_hours' not in columns:
        c.execute('ALTER TABLE attendance ADD COLUMN working_hours REAL DEFAULT 0.0')
        print("Added working_hours column to attendance table")
    
    if 'status' not in columns:
        c.execute('ALTER TABLE attendance ADD COLUMN status TEXT DEFAULT "present"')
        print("Added status column to attendance table")
    
    # Check if company_id column exists in users table
    c.execute("PRAGMA table_info(users)")
    user_columns = [column[1] for column in c.fetchall()]
    
    if 'company_id' not in user_columns:
        c.execute('ALTER TABLE users ADD COLUMN company_id INTEGER')
        print("Added company_id column to users table")
    
    if 'monthly_salary' not in user_columns:
        c.execute('ALTER TABLE users ADD COLUMN monthly_salary REAL DEFAULT 15000.0')
        print("Added monthly_salary column to users table")
    
    # Insert default companies
    c.execute('INSERT OR IGNORE INTO companies (name, description) VALUES (?, ?)', 
              ('Absolute Global Outsourcing', 'Main company'))
    c.execute('INSERT OR IGNORE INTO companies (name, description) VALUES (?, ?)', 
              ('Default Company', 'Default company for existing users'))
    
    # Update company_id for existing users
    c.execute('''UPDATE users SET company_id = (
        SELECT id FROM companies WHERE companies.name = users.company_name
    ) WHERE company_id IS NULL''')
    
    # Insert default holidays (2025)
    default_holidays = [
        ('2025-01-26', 'Republic Day', 'National Holiday'),
        ('2025-08-15', 'Independence Day', 'National Holiday'),
        ('2025-10-02', 'Gandhi Jayanti', 'National Holiday'),
        ('2025-12-25', 'Christmas Day', 'Public Holiday'),
        ('2025-01-01', 'New Year Day', 'Public Holiday'),
        ('2025-05-01', 'Labour Day', 'Public Holiday'),
    ]
    
    for holiday in default_holidays:
        c.execute('INSERT OR IGNORE INTO holidays (date, name, description) VALUES (?, ?, ?)', holiday)
    
    # Add some sample users if table is empty
    c.execute('SELECT COUNT(*) FROM users')
    if c.fetchone()[0] == 0:
        sample_users = [
            (1, 'Nikhil Sathe', 'Absolute Global Outsourcing', '04:00', '20:00', 'night', 8.0, 25000.0),
            (12, 'Akash Gaikwad', 'Absolute Global Outsourcing', '04:00', '19:00', 'night', 8.0, 20000.0),
            (13, 'Shubham Dabhane', 'Absolute Global Outsourcing', '19:00', '04:00', 'night', 8.0, 22000.0),
            (14, 'Shrutika Gaikwad', 'Absolute Global Outsourcing', '04:00', '20:00', 'night', 8.0, 18000.0),
            (15, 'Manoj Yadav', 'Absolute Global Outsourcing', '04:00', '20:00', 'night', 8.0, 21000.0),
            (41, 'Rohit Sahani', 'Absolute Global Outsourcing', '04:00', '19:00', 'night', 8.0, 23000.0)
        ]
        
        for user in sample_users:
            c.execute('''INSERT OR REPLACE INTO users 
                        (userid, name, company_name, shift_start_time, shift_end_time, shift_type, working_hours_per_day, monthly_salary) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)''', user)
        
        # Set company_id for sample users
        c.execute('''UPDATE users SET company_id = (
            SELECT id FROM companies WHERE companies.name = 'Absolute Global Outsourcing'
        )''')
        
        print("Added sample users to database")
    
    # Add sample attendance records if table is empty
    c.execute('SELECT COUNT(*) FROM attendance')
    if c.fetchone()[0] == 0:
        # Sample attendance records for demonstration
        sample_attendance = [
            # Today's records - Night shift pattern
            (1, '2025-08-17 19:30:00', '2025-08-17 19:30:00', '2025-08-18 04:30:00', 9.0),
            (12, '2025-08-17 19:00:00', '2025-08-17 19:00:00', '2025-08-18 04:00:00', 9.0),
            (13, '2025-08-17 19:30:00', '2025-08-17 19:30:00', '2025-08-18 04:30:00', 9.0),
            (14, '2025-08-17 19:00:00', '2025-08-17 19:00:00', '2025-08-18 04:00:00', 9.0),
            (15, '2025-08-17 19:30:00', '2025-08-17 19:30:00', '2025-08-18 04:30:00', 9.0),
            (41, '2025-08-17 19:00:00', '2025-08-17 19:00:00', '2025-08-18 04:00:00', 9.0),
            
            # Previous day examples - Night shift pattern
            (1, '2025-08-16 19:30:00', '2025-08-16 19:30:00', '2025-08-17 04:30:00', 9.0),
            (12, '2025-08-16 19:00:00', '2025-08-16 19:00:00', '2025-08-17 04:00:00', 9.0),
            (13, '2025-08-16 19:30:00', '2025-08-16 19:30:00', '2025-08-17 04:30:00', 9.0),
            (14, '2025-08-16 19:00:00', '2025-08-16 19:00:00', '2025-08-17 04:00:00', 9.0),
            (15, '2025-08-16 19:30:00', '2025-08-16 19:30:00', '2025-08-17 04:30:00', 9.0),
            (41, '2025-08-16 19:00:00', '2025-08-16 19:00:00', '2025-08-17 04:00:00', 9.0),
        ]
        
        for record in sample_attendance:
            c.execute('''INSERT INTO attendance 
                        (userid, timestamp, check_in, check_out, working_hours) 
                        VALUES (?, ?, ?, ?, ?)''', record)
        
        print("Added sample attendance records to database")
    
    conn.commit()
    conn.close()
    print("Database setup completed!")

def test_device_connection(ip, port=4370, timeout=5):
    """Test if a device is reachable at the given IP and port"""
    try:
        print(f"Testing connection to {ip}:{port}...")
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(timeout)
        result = sock.connect_ex((ip, port))
        sock.close()
        
        if result == 0:
            print(f"✓ Connection to {ip}:{port} successful")
            return True
        else:
            print(f"✗ Connection to {ip}:{port} failed (error code: {result})")
            return False
    except Exception as e:
        print(f"✗ Connection test error: {e}")
        return False

def pull_data_from_device(device_ip, device_port):
    """Pull users and attendance data from device"""
    try:
        # Test connection first
        if not test_device_connection(device_ip, device_port):
            return False, "Device connection failed"
        
        # Import ZK library
        try:
            from zk import ZK, const
        except ImportError:
            return False, "ZK library not available. Please install it with: pip install pyzk"
        
        # Connect to device
        print(f"Attempting to connect to ZK device at {device_ip}:{device_port}...")
        zk = ZK(device_ip, port=device_port, timeout=5)
        
        try:
            conn = zk.connect()
            if not conn:
                return False, f"Failed to establish ZK connection to {device_ip}:{device_port}"
            print(f"✓ ZK connection established successfully")
        except Exception as e:
            return False, f"ZK connection error: {str(e)}"
        
        try:
            # Get users
            users = conn.get_users()
            print(f"Found {len(users)} users on device")
            
            # Get attendance records
            attendance_records = conn.get_attendance()
            print(f"Found {len(attendance_records)} attendance records on device")
            
            # Process users
            db_conn = get_db_connection()
            cursor = db_conn.cursor()
            
            # Clear existing attendance only (preserve user data)
            cursor.execute('DELETE FROM attendance')
            
            # Insert or update users (preserve existing custom data)
            for user in users:
                # Check if user already exists
                existing_user = cursor.execute('SELECT * FROM users WHERE userid = ?', (user.user_id,)).fetchone()
                
                if existing_user:
                    # User exists - only update name if it changed, preserve all other custom data
                    if existing_user['name'] != user.name:
                        cursor.execute('UPDATE users SET name = ? WHERE userid = ?', (user.name, user.user_id))
                        print(f"Updated name for user {user.user_id}: {user.name}")
                else:
                    # New user - insert with default values
                    cursor.execute('''INSERT INTO users 
                        (userid, name, company_name, shift_start_time, shift_end_time, shift_type, working_hours_per_day, monthly_salary, created_date) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''', 
                        (user.user_id, user.name, 'Absolute Global Outsourcing', '09:00', '18:00', 'day', 8.0, 15000.0, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
                    print(f"Added new user: {user.name} (ID: {user.user_id})")
            
            # Process attendance records for night shift employees
            # For night shifts: early morning times are check-out, evening times are check-in
            night_shift_employees = set()  # Track which employees are night shifts
            
            for record in attendance_records:
                user_id = record.user_id
                timestamp = record.timestamp
                
                # Check if this user is a night shift employee
                user_conn = get_db_connection()
                user_info = user_conn.execute('SELECT shift_type FROM users WHERE userid = ?', (user_id,)).fetchone()
                user_conn.close()
                
                is_night_shift = user_info and user_info['shift_type'] == 'night' if user_info else False
                
                if is_night_shift:
                    night_shift_employees.add(user_id)
                
                # Insert raw attendance record
                cursor.execute('''INSERT INTO attendance 
                    (userid, timestamp, check_in, check_out, working_hours) 
                    VALUES (?, ?, ?, ?, ?)''', 
                    (user_id, timestamp.strftime('%Y-%m-%d %H:%M:%S'), 
                     timestamp.strftime('%Y-%m-%d %H:%M:%S'), None, 0.0))
            
            # Now process the attendance records to properly assign check-in/check-out
            # Group records by user and date, but handle multiple punches intelligently
            cursor.execute('''SELECT userid, DATE(timestamp) as date, timestamp
                             FROM attendance 
                             ORDER BY userid, DATE(timestamp), timestamp''')
            
            all_records = cursor.fetchall()
            
            # Group records by user and date
            user_date_punches = {}
            for record in all_records:
                user_id = record['userid']
                date = record['date']
                timestamp = record['timestamp']
                
                if user_id not in user_date_punches:
                    user_date_punches[user_id] = {}
                if date not in user_date_punches[user_id]:
                    user_date_punches[user_id][date] = []
                
                user_date_punches[user_id][date].append({'timestamp': timestamp})
            
            # Process each user's daily punches
            for user_id, dates in user_date_punches.items():
                for date, punches in dates.items():
                    # Use intelligent punch processing
                    check_in_time, check_out_time = process_multiple_punches(user_id, date, punches)
                    
                    if check_in_time and check_out_time:
                        # Update the attendance records with proper check-in/check-out times
                        cursor.execute('''UPDATE attendance 
                                         SET check_in = ?, check_out = ? 
                                         WHERE userid = ? AND DATE(timestamp) = ?''', 
                                      (check_in_time, check_out_time, user_id, date))
                        
                        # Calculate working hours
                        try:
                            check_in_dt = datetime.strptime(check_in_time, '%Y-%m-%d %H:%M:%S')
                            check_out_dt = datetime.strptime(check_out_time, '%Y-%m-%d %H:%M:%S')
                            
                            # Get user's shift information
                            user_conn = get_db_connection()
                            user_info = user_conn.execute('SELECT shift_type FROM users WHERE userid = ?', (user_id,)).fetchone()
                            user_conn.close()
                            
                            is_night_shift = user_info and user_info['shift_type'] == 'night' if user_info else False
                            
                            # For night shifts, if check-out is earlier than check-in, add 24 hours
                            if is_night_shift and check_out_dt < check_in_dt:
                                check_out_dt += timedelta(days=1)
                            
                            # Calculate working hours
                            time_diff = check_out_dt - check_in_dt
                            working_hours = time_diff.total_seconds() / 3600.0
                            
                            # Cap working hours at 10 for night shifts
                            if is_night_shift and working_hours > 10.0:
                                working_hours = 10.0
                            
                            # Update working hours
                            cursor.execute('''UPDATE attendance 
                                             SET working_hours = ? 
                                             WHERE userid = ? AND DATE(timestamp) = ?''', 
                                          (working_hours, user_id, date))
                            
                            print(f"Calculated working hours for user {user_id}: {working_hours:.2f} hours")
                            
                        except Exception as e:
                            print(f"Error calculating working hours for user {user_id}: {e}")
            
            db_conn.commit()
            db_conn.close()
            
            # Properly disconnect from device
            if conn and hasattr(conn, 'disconnect'):
                try:
                    conn.disconnect()
                except:
                    pass
            
            if zk and hasattr(zk, 'disconnect'):
                try:
                    zk.disconnect()
                except:
                    pass
            
            return True, f"Successfully pulled {len(users)} users and {len(attendance_records)} attendance records"
            
        except Exception as e:
            # Properly disconnect from device on error
            if conn and hasattr(conn, 'disconnect'):
                try:
                    conn.disconnect()
                except:
                    pass
            
            if zk and hasattr(zk, 'disconnect'):
                try:
                    zk.disconnect()
                except:
                    pass
            
            return False, f"Error processing device data: {str(e)}"
            
    except Exception as e:
        return False, f"Error connecting to device: {str(e)}"

def sync_users_from_device(device_ip, device_port):
    """Sync only user data from device (preserve existing custom data)"""
    try:
        # Test connection first
        if not test_device_connection(device_ip, device_port):
            return False, "Device connection failed"
        
        # Import ZK library
        try:
            from zk import ZK, const
        except ImportError:
            return False, "ZK library not available. Please install it with: pip install pyzk"
        
        # Connect to device
        zk = ZK(device_ip, port=device_port, timeout=5)
        conn = zk.connect()
        
        if not conn:
            return False, "Failed to connect to device"
        
        try:
            # Get users from device
            users = conn.get_users()
            print(f"Found {len(users)} users on device")
            
            # Process users
            db_conn = get_db_connection()
            cursor = db_conn.cursor()
            
            users_added = 0
            users_updated = 0
            
            # Insert or update users (preserve existing custom data)
            for user in users:
                # Check if user already exists
                existing_user = cursor.execute('SELECT * FROM users WHERE userid = ?', (user.user_id,)).fetchone()
                
                if existing_user:
                    # User exists - only update name if it changed, preserve all other custom data
                    if existing_user['name'] != user.name:
                        cursor.execute('UPDATE users SET name = ? WHERE userid = ?', (user.name, user.user_id))
                        print(f"Updated name for user {user.user_id}: {user.name}")
                        users_updated += 1
                else:
                    # New user - insert with default values
                    cursor.execute('''INSERT INTO users 
                        (userid, name, company_name, shift_start_time, shift_end_time, shift_type, working_hours_per_day, monthly_salary, created_date) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''', 
                        (user.user_id, user.name, 'Absolute Global Outsourcing', '09:00', '18:00', 'day', 8.0, 15000.0, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
                    print(f"Added new user: {user.name} (ID: {user.user_id})")
                    users_added += 1
            
            db_conn.commit()
            db_conn.close()
            
            # Properly disconnect from device
            if conn and hasattr(conn, 'disconnect'):
                try:
                    conn.disconnect()
                except:
                    pass
            
            if zk and hasattr(zk, 'disconnect'):
                try:
                    zk.disconnect()
                except:
                    pass
            
            return True, f"Successfully synced users: {users_added} added, {users_updated} updated"
            
        except Exception as e:
            # Properly disconnect from device on error
            if conn and hasattr(conn, 'disconnect'):
                try:
                    conn.disconnect()
                except:
                    pass
            
            if zk and hasattr(zk, 'disconnect'):
                try:
                    zk.disconnect()
                except:
                    pass
            
            return False, f"Error processing device data: {str(e)}"
            
    except Exception as e:
        return False, f"Error connecting to device: {str(e)}"

def pull_latest_data_from_device():
    """Pull latest data from the default device"""
    try:
        # Test connection first
        if not test_device_connection(DEFAULT_DEVICE_IP, DEFAULT_DEVICE_PORT):
            return False, f"Device at {DEFAULT_DEVICE_IP}:{DEFAULT_DEVICE_PORT} is not reachable"
        
        # Pull data
        success, message = pull_data_from_device(DEFAULT_DEVICE_IP, DEFAULT_DEVICE_PORT)
        return success, message
        
    except Exception as e:
        return False, f"Error pulling latest data: {str(e)}"

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    """Admin login page"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if verify_admin(username, password):
            session['admin_logged_in'] = True
            session['admin_username'] = username
            flash('Login successful!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password!', 'error')
    
    return render_template('admin_login.html', company_name=COMPANY_NAME, developer_name=DEVELOPER_NAME)

@app.route('/admin/logout')
def admin_logout():
    """Admin logout"""
    session.pop('admin_logged_in', None)
    session.pop('admin_username', None)
    flash('You have been logged out successfully!', 'success')
    return redirect(url_for('admin_login'))

@app.route('/')
def root():
    """Root route - redirect to admin login or dashboard"""
    if 'admin_logged_in' in session:
        return redirect(url_for('dashboard'))
    else:
        return redirect(url_for('admin_login'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Main dashboard page"""
    conn = get_db_connection()
    
    # Get total counts
    total_users = conn.execute('SELECT COUNT(*) FROM users').fetchone()[0]
    total_attendance = conn.execute('SELECT COUNT(*) FROM attendance').fetchone()[0]
    
    # Get today's attendance
    today = datetime.now().strftime('%Y-%m-%d')
    today_attendance = conn.execute('''SELECT COUNT(*) FROM attendance 
                                     WHERE DATE(timestamp) = ?''', (today,)).fetchone()[0]
    
    # Get recent attendance records
    recent_records = conn.execute('''SELECT u.name, a.timestamp, a.check_in 
                                   FROM attendance a 
                                   JOIN users u ON a.userid = u.userid 
                                   ORDER BY a.timestamp DESC LIMIT 10''').fetchall()
    
    # Check device status
    device_status = test_device_connection(DEFAULT_DEVICE_IP, DEFAULT_DEVICE_PORT)
    
    conn.close()
    
    return render_template('index.html', 
                         total_users=total_users,
                         total_attendance=total_attendance,
                         today_attendance=today_attendance,
                         recent_records=recent_records,
                         device_status=device_status,
                         device_ip=DEFAULT_DEVICE_IP,
                         device_port=DEFAULT_DEVICE_PORT,
                         company_name=COMPANY_NAME,
                         developer_name=DEVELOPER_NAME,
                         admin_username=session.get('admin_username'))

@app.route('/users')
@login_required
def users():
    """Users page"""
    conn = get_db_connection()
    users = conn.execute('''SELECT u.userid, u.name, u.company_name, u.shift_start_time, 
                           u.shift_end_time, u.shift_type, u.working_hours_per_day, u.monthly_salary, u.created_date,
                           c.name as company_display_name
                           FROM users u 
                           LEFT JOIN companies c ON u.company_id = c.id
                           ORDER BY u.name''').fetchall()
    
    # Get companies for the filter dropdown
    companies = conn.execute('SELECT * FROM companies ORDER BY name').fetchall()
    conn.close()
    
    return render_template('users.html', 
                         users=users, 
                         companies=companies,
                         company_name=COMPANY_NAME, 
                         developer_name=DEVELOPER_NAME, 
                         admin_username=session.get('admin_username'))

@app.route('/attendance')
@login_required
def attendance():
    """Attendance page"""
    # Import datetime at the function level to avoid scope issues
    from datetime import datetime, timedelta
    
    conn = get_db_connection()
    
    # Get filter parameters
    user_id = request.args.get('user_id', '')
    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')
    company_id = request.args.get('company_id', '')
    
    # Set default date range if not provided (last 30 days)
    if not from_date:
        from_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not to_date:
        to_date = datetime.now().strftime('%Y-%m-%d')
    
    # Get attendance records based on filters
    if user_id:
        records = conn.execute('''SELECT u.name, u.userid, a.timestamp, a.check_in, a.check_out, 
                                   a.working_hours, c.name as company_name
                                   FROM attendance a 
                                   JOIN users u ON a.userid = u.userid 
                                   LEFT JOIN companies c ON u.company_id = c.id
                                   WHERE DATE(a.timestamp) >= ? AND DATE(a.timestamp) <= ? AND u.userid = ?
                                   ORDER BY a.timestamp DESC''', (from_date, to_date, user_id)).fetchall()
    elif company_id:
        records = conn.execute('''SELECT u.name, u.userid, a.timestamp, a.check_in, a.check_out, 
                                   a.working_hours, c.name as company_name
                                   FROM attendance a 
                                   JOIN users u ON a.userid = u.userid 
                                   LEFT JOIN companies c ON u.company_id = c.id
                                   WHERE DATE(a.timestamp) >= ? AND DATE(a.timestamp) <= ? AND u.company_id = ?
                                   ORDER BY a.timestamp DESC''', (from_date, to_date, company_id)).fetchall()
    else:
        records = conn.execute('''SELECT u.name, u.userid, a.timestamp, a.check_in, a.check_out, 
                                   a.working_hours, c.name as company_name
                                   FROM attendance a 
                                   JOIN users u ON a.userid = u.userid 
                                   LEFT JOIN companies c ON u.company_id = c.id
                                   WHERE DATE(a.timestamp) >= ? AND DATE(a.timestamp) <= ?
                                   ORDER BY a.timestamp DESC''', (from_date, to_date)).fetchall()
    
    # Process records to add working hours and format data
    processed_records = []
    for record in records:
        record_dict = dict(record)
        
        # Use working hours from database if available, otherwise calculate
        if record_dict.get('working_hours') is not None and record_dict['working_hours'] > 0:
            # Use the stored working hours from database
            working_hours = record_dict['working_hours']
        else:
            # Calculate working hours if not available in database
            working_hours = 0.0
            
            if record_dict['check_in'] and record_dict['check_out']:
                try:
                    # Parse check-in and check-out times
                    check_in_str = str(record_dict['check_in'])
                    check_out_str = str(record_dict['check_out'])
                    
                    # Extract time part if it's a full datetime
                    if ' ' in check_in_str:
                        check_in_time = check_in_str.split(' ')[1]
                    else:
                        check_in_time = check_in_str
                        
                    if ' ' in check_out_str:
                        check_out_time = check_out_str.split(' ')[1]
                    else:
                        check_out_time = check_out_str
                    
                    # Check if check-in and check-out are the same time
                    if check_in_time == check_out_time:
                        # Same time means 0 working hours (single punch or incomplete record)
                        working_hours = 0.0
                    else:
                        # Parse times
                        check_in_parts = check_in_time.split(':')
                        check_out_parts = check_out_time.split(':')
                        
                        if len(check_in_parts) >= 2 and len(check_out_parts) >= 2:
                            try:
                                # Get user's shift information first
                                user_conn = get_db_connection()
                                user_shift = user_conn.execute('''SELECT shift_type, shift_start_time, shift_end_time 
                                                               FROM users WHERE userid = ?''', 
                                                            (record_dict['userid'],)).fetchone()
                                user_conn.close()
                                
                                # Parse time components
                                check_in_hour = int(check_in_parts[0])
                                check_in_minute = int(check_in_parts[1])
                                check_in_second = int(check_in_parts[2]) if len(check_in_parts) > 2 else 0
                                
                                check_out_hour = int(check_out_parts[0])
                                check_out_minute = int(check_out_parts[1])
                                check_out_second = int(check_out_parts[2]) if len(check_out_parts) > 2 else 0
                                
                                # For night shifts, the biometric device ALWAYS flips check-in and check-out
                                # So we need to ALWAYS swap them, not just when we detect a reversal
                                if user_shift and user_shift['shift_type'] == 'night':
                                    # For night shifts, ALWAYS swap the times:
                                    # What device shows as "check-in" (evening) is actually check-out
                                    # What device shows as "check-out" (morning) is actually check-in
                                    
                                    # Swap the times
                                    temp_hour = check_in_hour
                                    temp_minute = check_in_minute
                                    temp_second = check_in_second
                                    
                                    check_in_hour = check_out_hour
                                    check_in_minute = check_out_minute
                                    check_in_second = check_out_second
                                    
                                    check_out_hour = temp_hour
                                    check_out_minute = temp_minute
                                    check_out_second = temp_second
                                    
                                    print(f"Swapped times for night shift user {record_dict['userid']}: "
                                          f"Device check-in ({temp_hour:02d}:{temp_minute:02d}:{temp_second:02d}) → Real check-out, "
                                          f"Device check-out ({check_in_hour:02d}:{check_in_minute:02d}:{check_in_second:02d}) → Real check-in")
                                    
                                    # Now create proper datetime objects with correct dates
                                    # Check-in is on current day (morning, but actually evening after swap)
                                    check_in_date = datetime.now().date()
                                    # Check-out is on next day (evening, but actually morning after swap)
                                    check_out_date = check_in_date + timedelta(days=1)
                                    
                                    check_in_dt = datetime.combine(check_in_date, datetime.min.time().replace(
                                        hour=check_in_hour, minute=check_in_minute, second=check_in_second
                                    ))
                                    
                                    check_out_dt = datetime.combine(check_out_date, datetime.min.time().replace(
                                        hour=check_out_hour, minute=check_out_minute, second=check_out_second
                                    ))
                                    
                                    # For night shift, if check-out is earlier than check-in, add 1 day
                                    if check_out_dt < check_in_dt:
                                        check_out_dt += timedelta(days=1)
                                    
                                else:
                                    # For day shifts, use standard logic
                                    today = datetime.now().date()
                                    
                                    check_in_dt = datetime.combine(today, datetime.min.time().replace(
                                        hour=check_in_hour, minute=check_in_minute, second=check_in_second
                                    ))
                                    
                                    check_out_dt = datetime.combine(today, datetime.min.time().replace(
                                        hour=check_out_hour, minute=check_out_minute, second=check_out_second
                                    ))
                                    
                                    # For day shift, if check-out is earlier than check-in, it's likely next day
                                    if check_out_dt < check_in_dt:
                                        check_out_dt += timedelta(days=1)
                                
                                # Calculate the difference
                                time_diff = check_out_dt - check_in_dt
                                working_hours = time_diff.total_seconds() / 3600.0
                                
                                # Cap working hours at 10 for night shifts
                                if user_shift and user_shift['shift_type'] == 'night' and working_hours > 10.0:
                                    working_hours = 10.0
                                
                                # Format working hours to 2 decimal places
                                working_hours = round(working_hours, 2)
                                
                                # Store formatted time strings for display
                                record_dict['check_in_formatted'] = f"{check_in_hour:02d}:{check_in_minute:02d}:{check_in_second:02d}"
                                record_dict['check_out_formatted'] = f"{check_out_hour:02d}:{check_out_minute:02d}:{check_out_second:02d}"
                                
                            except Exception as e:
                                print(f"Error calculating working hours for user {record_dict['userid']}: {e}")
                                working_hours = 0.0
                                
                except Exception as e:
                    print(f"Error processing check-in/check-out times for user {record_dict['userid']}: {e}")
                    working_hours = 0.0
        
        # Add working hours to record
        record_dict['working_hours'] = working_hours
        
        # Format the record for display
        if record_dict['check_in']:
            if ' ' in str(record_dict['check_in']):
                record_dict['check_in_time'] = str(record_dict['check_in']).split(' ')[1]
            else:
                record_dict['check_in_time'] = str(record_dict['check_in'])
        else:
            record_dict['check_in_time'] = None
            
        if record_dict['check_out']:
            if ' ' in str(record_dict['check_out']):
                record_dict['check_out_time'] = str(record_dict['check_out']).split(' ')[1]
            else:
                record_dict['check_out_time'] = str(record_dict['check_out'])
        else:
            record_dict['check_out_time'] = None
        
        processed_records.append(record_dict)
    
    # Get all users for filter dropdown
    users = conn.execute('SELECT * FROM users ORDER BY name').fetchall()
    
    # Get all companies for filter dropdown
    companies = conn.execute('SELECT * FROM companies ORDER BY name').fetchall()
    
    # Get today's date for the template
    today_date = datetime.now().strftime('%Y-%m-%d')
    
    conn.close()
    
    return render_template('attendance.html', 
                         records=processed_records, 
                         users=users, 
                         companies=companies,
                         selected_user=user_id, 
                         from_date=from_date,
                         to_date=to_date,
                         selected_company=company_id,
                         today=today_date,
                         now=datetime.now(),
                         timedelta=timedelta,
                         company_name=COMPANY_NAME,
                         developer_name=DEVELOPER_NAME,
                         admin_username=session.get('admin_username'))

@app.route('/device')
@login_required
def device():
    """Device management page"""
    return render_template('device.html', company_name=COMPANY_NAME, developer_name=DEVELOPER_NAME, admin_username=session.get('admin_username'))

@app.route('/companies')
@login_required
def companies():
    """Companies management page"""
    conn = get_db_connection()
    companies_list = conn.execute('SELECT * FROM companies ORDER BY name').fetchall()
    conn.close()
    return render_template('companies.html', companies=companies_list, company_name=COMPANY_NAME, developer_name=DEVELOPER_NAME, admin_username=session.get('admin_username'))

@app.route('/api/test_connection', methods=['POST'])
def test_connection():
    """API endpoint to test device connection"""
    data = request.get_json()
    ip = data.get('ip')
    port = int(data.get('port', 4370))
    
    if test_device_connection(ip, port):
        return jsonify({'success': True, 'message': f'Device at {ip}:{port} is reachable!'})
    else:
        return jsonify({'success': False, 'message': f'Device at {ip}:{port} is not reachable!'})

@app.route('/api/pull_data', methods=['POST'])
def api_pull_data():
    """API endpoint to pull data from device"""
    try:
        data = request.get_json()
        device_ip = data.get('device_ip', DEFAULT_DEVICE_IP)
        device_port = data.get('device_port', DEFAULT_DEVICE_PORT)
        
        success, message = pull_data_from_device(device_ip, device_port)
        
        if success:
            return jsonify({'success': True, 'message': message})
        else:
            return jsonify({'success': False, 'message': message})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/pull_latest_data', methods=['POST'])
def pull_latest_data():
    """API endpoint to pull latest data from default device"""
    success, message = pull_latest_data_from_device()
    return jsonify({'success': success, 'message': message})

@app.route('/api/sync_users_only', methods=['POST'])
def sync_users_only():
    """API endpoint to sync only user data from device (preserve existing custom data)"""
    try:
        device_ip = request.form.get('device_ip', DEFAULT_DEVICE_IP)
        device_port = int(request.form.get('device_port', DEFAULT_DEVICE_PORT))
        
        success, message = sync_users_from_device(device_ip, device_port)
        
        if success:
            return jsonify({'success': True, 'message': message})
        else:
            return jsonify({'success': False, 'message': message})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/users', methods=['GET'])
def get_users():
    """API endpoint to get all users"""
    try:
        conn = get_db_connection()
        users = conn.execute('''SELECT u.userid, u.name, u.company_name, u.shift_start_time, 
                               u.shift_end_time, u.shift_type, u.working_hours_per_day, u.monthly_salary, u.created_date,
                               c.name as company_display_name
                               FROM users u 
                               LEFT JOIN companies c ON u.company_id = c.id
                               ORDER BY u.name''').fetchall()
        conn.close()
        
        users_list = []
        for user in users:
            users_list.append({
                'userid': user['userid'],
                'name': user['name'],
                'company_name': user['company_name'],
                'company_display_name': user['company_display_name'] or user['company_name'],
                'shift_start_time': user['shift_start_time'],
                'shift_end_time': user['shift_end_time'],
                'shift_type': user['shift_type'],
                'working_hours_per_day': user['working_hours_per_day'],
                'monthly_salary': user['monthly_salary'],
                'created_date': user['created_date']
            })
        
        return jsonify({'success': True, 'users': users_list})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/users/<int:userid>', methods=['GET'])
def get_user(userid):
    """API endpoint to get a specific user"""
    try:
        conn = get_db_connection()
        user = conn.execute('''SELECT u.userid, u.name, u.company_name, u.shift_start_time, 
                              u.shift_end_time, u.shift_type, u.working_hours_per_day, u.monthly_salary, u.created_date,
                              c.name as company_display_name
                              FROM users u 
                              LEFT JOIN companies c ON u.company_id = c.id
                              WHERE u.userid = ?''', (userid,)).fetchone()
        conn.close()
        
        if user:
            return jsonify({
                'success': True, 
                'user': {
                    'userid': user['userid'],
                    'name': user['name'],
                    'company_name': user['company_name'],
                    'company_display_name': user['company_display_name'] or user['company_name'],
                    'shift_start_time': user['shift_start_time'],
                    'shift_end_time': user['shift_end_time'],
                                    'shift_type': user['shift_type'],
                'working_hours_per_day': user['working_hours_per_day'],
                'monthly_salary': user['monthly_salary'],
                'created_date': user['created_date']
                }
            })
        else:
            return jsonify({'success': False, 'message': 'User not found'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/users/<int:userid>', methods=['PUT'])
def update_user(userid):
    """API endpoint to update a user"""
    try:
        data = request.get_json()
        name = data.get('name')
        company_name = data.get('company_name', 'Default Company')
        shift_start_time = data.get('shift_start_time', '09:00:00')
        shift_end_time = data.get('shift_end_time', '18:00:00')
        shift_type = data.get('shift_type', 'day')
        working_hours_per_day = data.get('working_hours_per_day', 8.0)
        monthly_salary = data.get('monthly_salary', 15000.0)
        
        if not name:
            return jsonify({'success': False, 'message': 'Name is required'})
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if user exists
        cursor.execute('SELECT userid FROM users WHERE userid = ?', (userid,))
        if not cursor.fetchone():
            conn.close()
            return jsonify({'success': False, 'message': 'User not found'})
        
        # Get or create company
        cursor.execute('SELECT id FROM companies WHERE name = ?', (company_name,))
        company = cursor.fetchone()
        
        if company:
            company_id = company['id']
        else:
            # Create new company if it doesn't exist
            cursor.execute('INSERT INTO companies (name, description) VALUES (?, ?)', 
                         (company_name, f'Company for user {name}'))
            company_id = cursor.lastrowid
        
        # Update user
        cursor.execute('''UPDATE users 
                         SET name = ?, company_name = ?, company_id = ?, shift_start_time = ?, 
                             shift_end_time = ?, shift_type = ?, working_hours_per_day = ?, monthly_salary = ?
                         WHERE userid = ?''', 
                      (name, company_name, company_id, shift_start_time, shift_end_time, 
                       shift_type, working_hours_per_day, monthly_salary, userid))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': f'User {name} updated successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/users/<int:userid>', methods=['DELETE'])
def delete_user(userid):
    """API endpoint to delete a user"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if user exists
        cursor.execute('SELECT name FROM users WHERE userid = ?', (userid,))
        user = cursor.fetchone()
        if not user:
            conn.close()
            return jsonify({'success': False, 'message': 'User not found'})
        
        # Delete user's attendance records first
        cursor.execute('DELETE FROM attendance WHERE userid = ?', (userid,))
        cursor.execute('DELETE FROM attendance_marking WHERE userid = ?', (userid,))
        cursor.execute('DELETE FROM salary_calculations WHERE userid = ?', (userid,))
        
        # Delete user
        cursor.execute('DELETE FROM users WHERE userid = ?', (userid,))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': f'User {user["name"]} deleted successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/users', methods=['POST'])
def create_user():
    """API endpoint to create a new user"""
    try:
        data = request.get_json()
        userid = data.get('userid')
        name = data.get('name')
        company_name = data.get('company_name', 'Default Company')
        shift_start_time = data.get('shift_start_time', '09:00:00')
        shift_end_time = data.get('shift_end_time', '18:00:00')
        shift_type = data.get('shift_type', 'day')
        working_hours_per_day = data.get('working_hours_per_day', 8.0)
        monthly_salary = data.get('monthly_salary', 15000.0)
        
        if not userid or not name:
            return jsonify({'success': False, 'message': 'User ID and Name are required'})
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if user ID already exists
        cursor.execute('SELECT userid FROM users WHERE userid = ?', (userid,))
        if cursor.fetchone():
            conn.close()
            return jsonify({'success': False, 'message': 'User ID already exists'})
        
        # Get or create company
        cursor.execute('SELECT id FROM companies WHERE name = ?', (company_name,))
        company = cursor.fetchone()
        
        if company:
            company_id = company['id']
        else:
            # Create new company if it doesn't exist
            cursor.execute('INSERT INTO companies (name, description) VALUES (?, ?)', 
                         (company_name, f'Company for user {name}'))
            company_id = cursor.lastrowid
        
        # Create user
        cursor.execute('''INSERT INTO users (userid, name, company_name, company_id, shift_start_time, 
                                           shift_end_time, shift_type, working_hours_per_day, monthly_salary)
                         VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''', 
                      (userid, name, company_name, company_id, shift_start_time, shift_end_time, 
                       shift_type, working_hours_per_day, monthly_salary))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': f'User {name} created successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/stats')
def get_stats():
    """API endpoint to get statistics"""
    conn = get_db_connection()
    
    total_users = conn.execute('SELECT COUNT(*) FROM users').fetchone()[0]
    total_attendance = conn.execute('SELECT COUNT(*) FROM attendance').fetchone()[0]
    
    # Get today's attendance
    today = datetime.now().strftime('%Y-%m-%d')
    today_attendance = conn.execute('''SELECT COUNT(*) FROM attendance 
                                     WHERE DATE(timestamp) = ?''', (today,)).fetchone()[0]
    
    conn.close()
    
    return jsonify({
        'total_users': total_users,
        'total_attendance': total_attendance,
        'today_attendance': today_attendance
    })

@app.route('/api/companies', methods=['GET'])
def get_companies():
    """API endpoint to get all companies"""
    try:
        conn = get_db_connection()
        companies = conn.execute('SELECT * FROM companies ORDER BY name').fetchall()
        conn.close()
        
        companies_list = []
        for company in companies:
            companies_list.append({
                'id': company['id'],
                'name': company['name'],
                'description': company['description'],
                'created_date': company['created_date']
            })
        
        return jsonify({'success': True, 'companies': companies_list})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/companies/<int:company_id>', methods=['GET'])
def get_company(company_id):
    """API endpoint to get a specific company"""
    try:
        conn = get_db_connection()
        company = conn.execute('SELECT * FROM companies WHERE id = ?', (company_id,)).fetchone()
        conn.close()
        
        if company:
            return jsonify({
                'success': True, 
                'company': {
                    'id': company['id'],
                    'name': company['name'],
                    'description': company['description'],
                    'created_date': company['created_date']
                }
            })
        else:
            return jsonify({'success': False, 'message': 'Company not found'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/companies/<int:company_id>', methods=['PUT'])
def update_company(company_id):
    """API endpoint to update a company"""
    try:
        data = request.get_json()
        name = data.get('name')
        description = data.get('description', '')
        
        if not name:
            return jsonify({'success': False, 'message': 'Company name is required'})
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if company exists
        cursor.execute('SELECT id FROM companies WHERE id = ?', (company_id,))
        if not cursor.fetchone():
            conn.close()
            return jsonify({'success': False, 'message': 'Company not found'})
        
        # Check if name already exists for another company
        cursor.execute('SELECT id FROM companies WHERE name = ? AND id != ?', (name, company_id))
        if cursor.fetchone():
            conn.close()
            return jsonify({'success': False, 'message': 'Company name already exists'})
        
        # Update company
        cursor.execute('''UPDATE companies 
                         SET name = ?, description = ?
                         WHERE id = ?''', 
                      (name, description, company_id))
        
        # Update users with this company
        cursor.execute('''UPDATE users 
                         SET company_name = ?
                         WHERE company_id = ?''', (name, company_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': f'Company {name} updated successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/companies/<int:company_id>', methods=['DELETE'])
def delete_company(company_id):
    """API endpoint to delete a company"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if company exists
        cursor.execute('SELECT name FROM companies WHERE id = ?', (company_id,))
        company = cursor.fetchone()
        if not company:
            conn.close()
            return jsonify({'success': False, 'message': 'Company not found'})
        
        # Check if any users are using this company
        cursor.execute('SELECT COUNT(*) FROM users WHERE company_id = ?', (company_id,))
        user_count = cursor.fetchone()[0]
        
        if user_count > 0:
            conn.close()
            return jsonify({'success': False, 'message': f'Cannot delete company. {user_count} users are still assigned to this company.'})
        
        # Delete company
        cursor.execute('DELETE FROM companies WHERE id = ?', (company_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': f'Company {company["name"]} deleted successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/companies', methods=['POST'])
def create_company():
    """API endpoint to create a new company"""
    try:
        data = request.get_json()
        name = data.get('name')
        description = data.get('description', '')
        
        if not name:
            return jsonify({'success': False, 'message': 'Company name is required'})
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if company name already exists
        cursor.execute('SELECT id FROM companies WHERE name = ?', (name,))
        if cursor.fetchone():
            conn.close()
            return jsonify({'success': False, 'message': 'Company name already exists'})
        
        # Create company
        cursor.execute('''INSERT INTO companies (name, description)
                         VALUES (?, ?)''', 
                      (name, description))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': f'Company {name} created successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/recalculate_working_hours', methods=['POST'])
def recalculate_working_hours():
    """Recalculate working hours for all attendance records"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get all attendance records with check-in and check-out times
        records = cursor.execute('''SELECT a.id, a.userid, a.check_in, a.check_out, u.shift_type
                                   FROM attendance a 
                                   JOIN users u ON a.userid = u.userid 
                                   WHERE a.check_in IS NOT NULL AND a.check_out IS NOT NULL''').fetchall()
        
        updated_count = 0
        
        for record in records:
            try:
                check_in_str = str(record['check_in'])
                check_out_str = str(record['check_out'])
                shift_type = record['shift_type']
                
                # Parse times
                if ' ' in check_in_str:
                    check_in_time = check_in_str.split(' ')[1]
                else:
                    check_in_time = check_in_str
                    
                if ' ' in check_out_str:
                    check_out_time = check_out_str.split(' ')[1]
                else:
                    check_out_time = check_out_str
                
                # Check if check-in and check-out are the same time
                if check_in_time == check_out_time:
                    working_hours = 0.0
                else:
                    # Parse time components
                    check_in_parts = check_in_time.split(':')
                    check_out_parts = check_out_time.split(':')
                    
                    if len(check_in_parts) >= 2 and len(check_out_parts) >= 2:
                        # Create datetime objects for calculation
                        today = datetime.now().date()
                        
                        check_in_hour = int(check_in_parts[0])
                        check_in_minute = int(check_in_parts[1])
                        check_in_second = int(check_in_parts[2]) if len(check_in_parts) > 2 else 0
                        
                        check_out_hour = int(check_out_parts[0])
                        check_out_minute = int(check_out_parts[1])
                        check_out_second = int(check_out_parts[2]) if len(check_out_parts) > 2 else 0
                        
                        check_in_dt = datetime.combine(today, datetime.min.time().replace(
                            hour=check_in_hour, minute=check_in_minute, second=check_in_second
                        ))
                        
                        check_out_dt = datetime.combine(today, datetime.min.time().replace(
                            hour=check_out_hour, minute=check_out_minute, second=check_out_second
                        ))
                        
                        # Handle night shift logic
                        if shift_type == 'night':
                            # For night shifts, the biometric device ALWAYS flips check-in and check-out
                            # So we need to ALWAYS swap them, not just when we detect a reversal
                            
                            # Swap the times
                            temp_hour = check_in_hour
                            temp_minute = check_in_minute
                            temp_second = check_in_second
                            
                            check_in_hour = check_out_hour
                            check_in_minute = check_out_minute
                            check_in_second = check_out_second
                            
                            check_out_hour = temp_hour
                            check_out_minute = temp_minute
                            check_out_second = temp_second
                            
                            print(f"Swapped times for night shift user in recalc: "
                                  f"Device check-in ({temp_hour:02d}:{temp_minute:02d}:{temp_second:02d}) → Real check-out, "
                                  f"Device check-out ({check_in_hour:02d}:{check_in_minute:02d}:{check_in_second:02d}) → Real check-in")
                            
                            # Now create proper datetime objects with correct dates
                            # Check-in is on current day (morning, but actually evening after swap)
                            check_in_date = datetime.now().date()
                            # Check-out is on next day (evening, but actually morning after swap)
                            check_out_date = check_in_date + timedelta(days=1)
                            
                            check_in_dt = datetime.combine(check_in_date, datetime.min.time().replace(
                                hour=check_in_hour, minute=check_in_minute, second=check_in_second
                            ))
                            
                            check_out_dt = datetime.combine(check_out_date, datetime.min.time().replace(
                                hour=check_out_hour, minute=check_out_minute, second=check_out_second
                            ))
                            
                            # For night shift, if check-out is earlier than check-in, add 1 day
                            if check_out_dt < check_in_dt:
                                check_out_dt += timedelta(days=1)
                        else:
                            # For day shifts, if check-out is earlier than check-in, it's likely next day
                            if check_out_dt < check_in_dt:
                                check_out_dt += timedelta(days=1)
                        
                        # Calculate working hours
                        time_diff = check_out_dt - check_in_dt
                        working_hours = time_diff.total_seconds() / 3600.0
                        
                        # Cap working hours at 10 for night shifts
                        if shift_type == 'night' and working_hours > 10.0:
                            working_hours = 10.0
                        
                        # Format working hours to 2 decimal places
                        working_hours = round(working_hours, 2)
                
                # Update working hours in database
                cursor.execute('UPDATE attendance SET working_hours = ? WHERE id = ?', 
                             (working_hours, record['id']))
                updated_count += 1
                
            except Exception as e:
                print(f"Error processing record {record['id']}: {e}")
                continue
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True, 
            'message': f'Successfully recalculated working hours for {updated_count} records'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/add_sample_data', methods=['POST'])
def add_sample_data():
    """Add sample attendance data for testing"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Sample attendance records for demonstration
        sample_attendance = [
            # Today's records - Night shift pattern
            (1, '2025-08-17 19:30:00', '2025-08-17 19:30:00', '2025-08-18 04:30:00', 9.0),
            (12, '2025-08-17 19:00:00', '2025-08-17 19:00:00', '2025-08-18 04:00:00', 9.0),
            (13, '2025-08-17 19:30:00', '2025-08-17 19:30:00', '2025-08-18 04:30:00', 9.0),
            (14, '2025-08-17 19:00:00', '2025-08-17 19:00:00', '2025-08-18 04:00:00', 9.0),
            (15, '2025-08-17 19:30:00', '2025-08-17 19:30:00', '2025-08-18 04:30:00', 9.0),
            (41, '2025-08-17 19:00:00', '2025-08-17 19:00:00', '2025-08-18 04:00:00', 9.0),
            
            # Previous day examples - Night shift pattern
            (1, '2025-08-16 19:30:00', '2025-08-16 19:30:00', '2025-08-17 04:30:00', 9.0),
            (12, '2025-08-16 19:00:00', '2025-08-16 19:00:00', '2025-08-17 04:00:00', 9.0),
            (13, '2025-08-16 19:30:00', '2025-08-16 19:30:00', '2025-08-17 04:30:00', 9.0),
            (14, '2025-08-16 19:00:00', '2025-08-16 19:00:00', '2025-08-17 04:00:00', 9.0),
            (15, '2025-08-16 19:30:00', '2025-08-16 19:30:00', '2025-08-17 04:30:00', 9.0),
            (41, '2025-08-16 19:00:00', '2025-08-16 19:00:00', '2025-08-17 04:00:00', 9.0),
        ]
        
        # Clear existing attendance records first
        cursor.execute('DELETE FROM attendance')
        
        # Insert new sample records
        for record in sample_attendance:
            cursor.execute('''INSERT INTO attendance 
                            (userid, timestamp, check_in, check_out, working_hours) 
                            VALUES (?, ?, ?, ?, ?)''', record)
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True, 
            'message': f'Added {len(sample_attendance)} sample attendance records'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/clear_sample_data', methods=['POST'])
def clear_sample_data():
    """Clear sample attendance data"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Clear all attendance records
        cursor.execute('DELETE FROM attendance')
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True, 
            'message': 'All attendance records cleared successfully'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/push_data_to_device', methods=['POST'])
def push_data_to_device():
    """Push updated user data back to the biometric device"""
    try:
        data = request.get_json()
        device_ip = data.get('device_ip', DEFAULT_DEVICE_IP)
        device_port = data.get('device_port', DEFAULT_DEVICE_PORT)
        
        # Test connection first
        if not test_device_connection(device_ip, device_port):
            return jsonify({'success': False, 'message': f"Device at {device_ip}:{device_port} is not reachable"})
        
        # Import ZK library
        try:
            from zk import ZK, const
        except ImportError:
            return jsonify({'success': False, 'message': "ZK library not available. Please install it with: pip install pyzk"})
        
        # Connect to device
        print(f"Attempting to connect to ZK device at {device_ip}:{device_port}...")
        zk = ZK(device_ip, port=device_port, timeout=5)
        
        try:
            conn = zk.connect()
            if not conn:
                return jsonify({'success': False, 'message': f"Failed to establish ZK connection to {device_ip}:{device_port}"})
            print(f"✓ ZK connection established successfully")
        except Exception as e:
            return jsonify({'success': False, 'message': f"ZK connection error: {str(e)}"})
        
        try:
            # Get all users from database
            db_conn = get_db_connection()
            cursor = db_conn.cursor()
            cursor.execute('''SELECT userid, name, company_name, shift_start_time, shift_end_time, 
                                    shift_type, working_hours_per_day, monthly_salary 
                             FROM users ORDER BY userid''')
            users = cursor.fetchall()
            db_conn.close()
            
            if not users:
                return jsonify({'success': False, 'message': 'No users found in database'})
            
            # Push users to device
            users_updated = 0
            users_added = 0
            
            for user in users:
                try:
                    # Check if user exists on device
                    device_users = conn.get_users()
                    existing_device_user = None
                    
                    for device_user in device_users:
                        if device_user.user_id == user['userid']:
                            existing_device_user = device_user
                            break
                    
                    if existing_device_user:
                        # Update existing user on device
                        if existing_device_user.name != user['name']:
                            # Update name on device
                            conn.set_user(uid=user['userid'], name=user['name'])
                            print(f"Updated user {user['userid']} name on device: {user['name']}")
                            users_updated += 1
                    else:
                        # Add new user to device
                        conn.set_user(uid=user['userid'], name=user['name'])
                        print(f"Added user {user['userid']} to device: {user['name']}")
                        users_added += 1
                        
                except Exception as e:
                    print(f"Error processing user {user['userid']}: {e}")
                    continue
            
            # Properly disconnect from device
            if conn and hasattr(conn, 'disconnect'):
                try:
                    conn.disconnect()
                except:
                    pass
            
            if zk and hasattr(zk, 'disconnect'):
                try:
                    zk.disconnect()
                except:
                    pass
            
            return jsonify({
                'success': True, 
                'message': f'Successfully pushed data to device: {users_updated} updated, {users_added} added'
            })
            
        except Exception as e:
            # Properly disconnect from device on error
            if conn and hasattr(conn, 'disconnect'):
                try:
                    conn.disconnect()
                except:
                    pass
            
            if zk and hasattr(zk, 'disconnect'):
                try:
                    zk.disconnect()
                except:
                    pass
            
            return jsonify({'success': False, 'message': f'Error pushing data to device: {str(e)}'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/push_user_to_device', methods=['POST'])
def push_user_to_device():
    """Push a specific user's updated data to the biometric device"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        device_ip = data.get('device_ip', DEFAULT_DEVICE_IP)
        device_port = data.get('device_port', DEFAULT_DEVICE_PORT)
        
        if not user_id:
            return jsonify({'success': False, 'message': 'User ID is required'})
        
        # Test connection first
        if not test_device_connection(device_ip, device_port):
            return jsonify({'success': False, 'message': f"Device at {device_ip}:{device_port} is not reachable"})
        
        # Import ZK library
        try:
            from zk import ZK, const
        except ImportError:
            return jsonify({'success': False, 'message': "ZK library not available. Please install it with: pip install pyzk"})
        
        # Connect to device
        print(f"Attempting to connect to ZK device at {device_ip}:{device_port}...")
        zk = ZK(device_ip, port=device_port, timeout=5)
        
        try:
            conn = zk.connect()
            if not conn:
                return jsonify({'success': False, 'message': f"Failed to establish ZK connection to {device_ip}:{device_port}"})
            print(f"✓ ZK connection established successfully")
        except Exception as e:
            return jsonify({'success': False, 'message': f"ZK connection error: {str(e)}"})
        
        try:
            # Get user data from database
            db_conn = get_db_connection()
            cursor = db_conn.cursor()
            cursor.execute('''SELECT userid, name, company_name, shift_start_time, shift_end_time, 
                                    shift_type, working_hours_per_day, monthly_salary 
                             FROM users WHERE userid = ?''', (user_id,))
            user = cursor.fetchone()
            db_conn.close()
            
            if not user:
                return jsonify({'success': False, 'message': f'User {user_id} not found in database'})
            
            # Check if user exists on device
            device_users = conn.get_users()
            existing_device_user = None
            
            for device_user in device_users:
                if device_user.user_id == user['userid']:
                    existing_device_user = device_user
                    break
            
            if existing_device_user:
                # Update existing user on device
                if existing_device_user.name != user['name']:
                    conn.set_user(uid=user['userid'], name=user['name'])
                    print(f"Updated user {user['userid']} name on device: {user['name']}")
                    message = f"Updated user {user['name']} on device"
                else:
                    message = f"User {user['name']} already up to date on device"
            else:
                # Add new user to device
                conn.set_user(uid=user['userid'], name=user['name'])
                print(f"Added user {user['userid']} to device: {user['name']}")
                message = f"Added user {user['name']} to device"
            
            # Properly disconnect from device
            if conn and hasattr(conn, 'disconnect'):
                try:
                    conn.disconnect()
                except:
                    pass
            
            if zk and hasattr(zk, 'disconnect'):
                try:
                    zk.disconnect()
                except:
                    pass
            
            return jsonify({
                'success': True, 
                'message': message
            })
            
        except Exception as e:
            # Properly disconnect from device on error
            if conn and hasattr(conn, 'disconnect'):
                try:
                    conn.disconnect()
                except:
                    pass
            
            if zk and hasattr(zk, 'disconnect'):
                try:
                    zk.disconnect()
                except:
                    pass
            
            return jsonify({'success': False, 'message': f'Error pushing user to device: {str(e)}'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/test_reversed_times', methods=['POST'])
def test_reversed_times():
    """Test the night shift time swapping logic"""
    try:
        from datetime import datetime, timedelta
        
        # Test case: Biometric device shows flipped times for night shift
        # Device shows: Check-in: 04:30 (morning), Check-out: 19:30 (evening)
        # Reality: Check-in: 19:30 (evening), Check-out: 04:30 (next morning)
        
        device_check_in_hour = 4   # 04:30 AM (what device shows as check-in)
        device_check_in_minute = 30
        device_check_out_hour = 19 # 19:30 PM (what device shows as check-out)
        device_check_out_minute = 30
        
        # For night shifts, ALWAYS swap the times
        # Swap the times
        temp_hour = device_check_in_hour
        temp_minute = device_check_in_minute
        
        real_check_in_hour = device_check_out_hour
        real_check_in_minute = device_check_out_minute
        
        real_check_out_hour = temp_hour
        real_check_out_minute = temp_minute
        
        # Create proper datetime objects with correct dates
        check_in_date = datetime.now().date()
        check_out_date = check_in_date + timedelta(days=1)
        
        real_check_in_dt = datetime.combine(check_in_date, datetime.min.time().replace(
            hour=real_check_in_hour, minute=real_check_in_minute, second=0
        ))
        
        real_check_out_dt = datetime.combine(check_out_date, datetime.min.time().replace(
            hour=real_check_out_hour, minute=real_check_out_minute, second=0
        ))
        
        # For night shift, if check-out is earlier than check-in, add 1 day
        if real_check_out_dt < real_check_in_dt:
            real_check_out_dt += timedelta(days=1)
        
        # Calculate working hours
        time_diff = real_check_out_dt - real_check_in_dt
        working_hours = time_diff.total_seconds() / 3600.0
        
        result = {
            'success': True,
            'device_shows': {
                'check_in': f"{device_check_in_hour:02d}:{device_check_in_minute:02d}",
                'check_out': f"{device_check_out_hour:02d}:{device_check_out_minute:02d}"
            },
            'times_swapped': True,
            'corrected_times': {
                'check_in': f"{real_check_in_hour:02d}:{real_check_in_minute:02d}",
                'check_out': f"{real_check_out_hour:02d}:{real_check_out_minute:02d}"
            },
            'working_hours': round(working_hours, 2),
            'explanation': 'For night shifts, device times are ALWAYS swapped. Morning time (04:30) becomes check-out, evening time (19:30) becomes check-in.'
        }
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

def process_multiple_punches(user_id, date, punches):
    """Process multiple punches for a user on a given date to determine check-in/check-out"""
    try:
        # Get user's shift information
        user_conn = get_db_connection()
        user_info = user_conn.execute('SELECT shift_type FROM users WHERE userid = ?', (user_id,)).fetchone()
        user_conn.close()
        
        is_night_shift = user_info and user_info['shift_type'] == 'night' if user_info else False
        
        if not punches:
            return None, None
        
        # Sort punches by timestamp
        sorted_punches = sorted(punches, key=lambda x: x['timestamp'])
        
        if len(sorted_punches) == 1:
            # Single punch - treat as check-in only
            return sorted_punches[0]['timestamp'], None
        
        if is_night_shift:
            # For night shifts, biometric device ALWAYS flips times
            # First punch (morning) = check-out, last punch (evening) = check-in
            
            # Find the earliest morning punch (before 12:00) and latest evening punch (after 12:00)
            morning_punches = [p for p in sorted_punches if int(p['timestamp'].split(' ')[1].split(':')[0]) < 12]
            evening_punches = [p for p in sorted_punches if int(p['timestamp'].split(' ')[1].split(':')[0]) >= 12]
            
            if morning_punches and evening_punches:
                # Use earliest morning punch as check-out, latest evening punch as check-in
                check_out_time = min(morning_punches, key=lambda x: x['timestamp'])
                check_in_time = max(evening_punches, key=lambda x: x['timestamp'])
                
                print(f"Night shift multiple punches for user {user_id}: "
                      f"Check-in: {check_in_time['timestamp']} (evening), "
                      f"Check-out: {check_out_time['timestamp']} (morning)")
                
                return check_in_time['timestamp'], check_out_time['timestamp']
            else:
                # Fallback: use first and last punches
                return sorted_punches[-1]['timestamp'], sorted_punches[0]['timestamp']
        else:
            # For day shifts, use standard logic
            # First punch = check-in, last punch = check-out
            check_in_time = sorted_punches[0]['timestamp']
            check_out_time = sorted_punches[-1]['timestamp']
            
            print(f"Day shift multiple punches for user {user_id}: "
                  f"Check-in: {check_in_time}, Check-out: {check_out_time}")
            
            return check_in_time, check_out_time
            
    except Exception as e:
        print(f"Error processing multiple punches for user {user_id}: {e}")
        return None, None

@app.route('/holidays')
@login_required
def holidays():
    """Holidays management page"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get all holidays
    cursor.execute('SELECT * FROM holidays ORDER BY date DESC')
    holidays = cursor.fetchall()
    
    conn.close()
    return render_template('holidays.html', holidays=holidays)

@app.route('/attendance_marking')
@login_required
def attendance_marking():
    """Attendance marking page"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get date filter
    date_filter = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    company_filter = request.args.get('company_id', '')
    
    # Build query
    query = '''
        SELECT u.userid, u.name, u.company_name, u.shift_start_time, u.shift_end_time, 
               u.shift_type, u.working_hours_per_day, u.monthly_salary,
               am.status as marked_status, am.working_hours as marked_hours,
               am.overtime_hours, am.late_minutes, am.remarks,
               a.check_in, a.check_out, a.working_hours as actual_hours
        FROM users u
        LEFT JOIN attendance_marking am ON u.userid = am.userid AND am.date = ?
        LEFT JOIN attendance a ON u.userid = a.userid AND DATE(a.timestamp) = ?
    '''
    
    params = [date_filter, date_filter]
    
    if company_filter:
        query += ' WHERE u.company_id = ?'
        params.append(company_filter)
    
    query += ' ORDER BY u.userid'
    
    cursor.execute(query, params)
    users = cursor.fetchall()
    
    # Get companies for filter
    cursor.execute('SELECT * FROM companies ORDER BY name')
    companies = cursor.fetchall()
    
    # Check if date is weekend or holiday
    date_obj = datetime.strptime(date_filter, '%Y-%m-%d')
    is_weekend = date_obj.weekday() >= 5  # Saturday = 5, Sunday = 6
    
    cursor.execute('SELECT * FROM holidays WHERE date = ?', (date_filter,))
    holiday = cursor.fetchone()
    
    conn.close()
    
    return render_template('attendance_marking.html', 
                         users=users, 
                         companies=companies,
                         selected_date=date_filter,
                         selected_company=company_filter,
                         is_weekend=is_weekend,
                         holiday=holiday)

@app.route('/salary')
@login_required
def salary():
    """Salary calculation page"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get filters
    month_filter = request.args.get('month', datetime.now().strftime('%B'))
    year_filter = request.args.get('year', datetime.now().year)
    company_filter = request.args.get('company_id', '')
    
    # Build query
    query = '''
        SELECT u.userid, u.name, u.company_name, u.monthly_salary,
               sc.total_days, sc.present_days, sc.absent_days, sc.leave_days,
               sc.total_working_hours, sc.overtime_hours,
               sc.basic_salary, sc.overtime_pay, sc.deductions, sc.net_salary
        FROM users u
        LEFT JOIN salary_calculations sc ON u.userid = sc.userid 
            AND sc.month = ? AND sc.year = ?
    '''
    
    params = [month_filter, year_filter]
    
    if company_filter:
        query += ' WHERE u.company_id = ?'
        params.append(company_filter)
    
    query += ' ORDER BY u.userid'
    
    cursor.execute(query, params)
    users = cursor.fetchall()
    
    # Get companies for filter
    cursor.execute('SELECT * FROM companies ORDER BY name')
    companies = cursor.fetchall()
    
    # Get available months and years
    cursor.execute('SELECT DISTINCT month, year FROM salary_calculations ORDER BY year DESC, month')
    available_periods = cursor.fetchall()
    
    conn.close()
    
    return render_template('salary.html', 
                         users=users, 
                         companies=companies,
                         selected_month=month_filter,
                         selected_year=year_filter,
                         selected_company=company_filter,
                         available_periods=available_periods)

@app.route('/email_management')
@login_required
def email_management():
    """Email management page"""
    return render_template('email_management.html')

@app.route('/analytics')
@login_required
def analytics():
    """Analytics dashboard page"""
    from datetime import datetime, timedelta
    now = datetime.now()
    return render_template('analytics.html', now=now, timedelta=timedelta)

@app.route('/api/attendance_marking', methods=['POST'])
# @login_required  # Temporarily commented out for debugging
def api_attendance_marking():
    """Mark attendance for a user"""
    try:
        data = request.get_json()
        print(f"Received data: {data}")  # Debug log
        
        userid = data.get('userid')
        date = data.get('date')
        status = data.get('status')
        working_hours = data.get('working_hours', 0.0)
        overtime_hours = data.get('overtime_hours', 0.0)
        late_minutes = data.get('late_minutes', 0)
        remarks = data.get('remarks', '')
        
        print(f"Parsed data - userid: {userid}, date: {date}, status: {status}")  # Debug log
        
        if not userid or not date or not status:
            return jsonify({'success': False, 'message': 'Missing required fields'})
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if marking already exists
        cursor.execute('SELECT * FROM attendance_marking WHERE userid = ? AND date = ?', (userid, date))
        existing = cursor.fetchone()
        
        print(f"Existing marking: {existing}")  # Debug log
        
        # Also check what's in the table for this user
        cursor.execute('SELECT * FROM attendance_marking WHERE userid = ? ORDER BY date', (userid,))
        all_markings = cursor.fetchall()
        print(f"All markings for user {userid}: {len(all_markings)} records")
        for marking in all_markings[:5]:  # Show first 5
            print(f"  {marking['date']}: {marking['status']}")
        
        if existing:
            # Update existing marking
            print(f"Updating existing marking for user {userid} on {date}")  # Debug log
            cursor.execute('''UPDATE attendance_marking 
                             SET status = ?, working_hours = ?, overtime_hours = ?, 
                                 late_minutes = ?, remarks = ?, marked_by = ?
                             WHERE userid = ? AND date = ?''', 
                          (status, working_hours, overtime_hours, late_minutes, remarks, 'admin', userid, date))
        else:
            # Create new marking
            print(f"Creating new marking for user {userid} on {date}")  # Debug log
            cursor.execute('''INSERT INTO attendance_marking 
                             (userid, date, status, working_hours, overtime_hours, late_minutes, remarks, marked_by)
                             VALUES (?, ?, ?, ?, ?, ?, ?, ?)''', 
                          (userid, date, status, working_hours, overtime_hours, late_minutes, remarks, 'admin'))
        
        conn.commit()
        print(f"Database committed successfully")  # Debug log
        
        # Verify the data was saved
        cursor.execute('SELECT * FROM attendance_marking WHERE userid = ? AND date = ?', (userid, date))
        saved_data = cursor.fetchone()
        print(f"Verified saved data: {saved_data}")  # Debug log
        
        conn.close()
        
        return jsonify({'success': True, 'message': 'Attendance marked successfully'})
        
    except Exception as e:
        print(f"Error in attendance_marking API: {str(e)}")  # Debug log
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/holidays', methods=['GET', 'POST'])
@login_required
def api_holidays():
    """Get or create holidays"""
    if request.method == 'GET':
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM holidays ORDER BY date DESC')
        holidays = cursor.fetchall()
        conn.close()
        
        # Convert to list of dictionaries
        holidays_list = []
        for holiday in holidays:
            holidays_list.append({
                'id': holiday['id'],
                'date': holiday['date'],
                'name': holiday['name'],
                'description': holiday['description'],
                'is_public_holiday': holiday['is_public_holiday']
            })
        
        return jsonify({'success': True, 'holidays': holidays_list})
    
    elif request.method == 'POST':
        try:
            data = request.get_json()
            date = data.get('date')
            name = data.get('name')
            description = data.get('description', '')
            is_public_holiday = data.get('is_public_holiday', True)
            
            if not date or not name:
                return jsonify({'success': False, 'message': 'Missing required fields'})
            
            conn = get_db_connection()
            cursor = conn.cursor()
            
            cursor.execute('''INSERT INTO holidays (date, name, description, is_public_holiday)
                             VALUES (?, ?, ?, ?)''', (date, name, description, is_public_holiday))
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True, 'message': 'Holiday added successfully'})
            
        except Exception as e:
            return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/holidays/<int:holiday_id>', methods=['PUT', 'DELETE'])
@login_required
def api_holiday_management(holiday_id):
    """Update or delete a holiday"""
    if request.method == 'PUT':
        try:
            data = request.get_json()
            date = data.get('date')
            name = data.get('name')
            description = data.get('description', '')
            is_public_holiday = data.get('is_public_holiday', True)
            
            if not date or not name:
                return jsonify({'success': False, 'message': 'Missing required fields'})
            
            conn = get_db_connection()
            cursor = conn.cursor()
            
            cursor.execute('''UPDATE holidays 
                             SET date = ?, name = ?, description = ?, is_public_holiday = ?
                             WHERE id = ?''', (date, name, description, is_public_holiday, holiday_id))
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True, 'message': 'Holiday updated successfully'})
            
        except Exception as e:
            return jsonify({'success': False, 'message': f'Error: {str(e)}'})
    
    elif request.method == 'DELETE':
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            cursor.execute('DELETE FROM holidays WHERE id = ?', (holiday_id,))
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True, 'message': 'Holiday deleted successfully'})
            
        except Exception as e:
            return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/calculate_salary', methods=['POST'])
@login_required
def api_calculate_salary():
    """Calculate salary for all users for a given month"""
    try:
        data = request.get_json()
        month = data.get('month')
        year = data.get('year')
        
        if not month or not year:
            return jsonify({'success': False, 'message': 'Missing month or year'})
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get all users
        cursor.execute('SELECT userid, monthly_salary, working_hours_per_day FROM users')
        users = cursor.fetchall()
        
        # Calculate working days in month
        start_date = datetime(int(year), datetime.strptime(month, '%B').month, 1)
        if start_date.month == 12:
            end_date = datetime(int(year) + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = datetime(int(year), start_date.month + 1, 1) - timedelta(days=1)
        
        working_days = 0
        current_date = start_date
        while current_date <= end_date:
            # Skip weekends
            if current_date.weekday() < 5:  # Monday to Friday
                # Check if it's a holiday
                cursor.execute('SELECT * FROM holidays WHERE date = ?', (current_date.strftime('%Y-%m-%d'),))
                if not cursor.fetchone():
                    working_days += 1
            current_date += timedelta(days=1)
        
        # Process each user
        for user in users:
            userid = user['userid']
            monthly_salary = user['monthly_salary']
            daily_hours = user['working_hours_per_day']
            
            # Get attendance marking for the month
            cursor.execute('''SELECT status, working_hours, overtime_hours, late_minutes
                             FROM attendance_marking 
                             WHERE userid = ? AND strftime('%Y-%m', date) = ?''', 
                          (userid, f"{year:04d}-{datetime.strptime(month, '%B').month:02d}"))
            
            monthly_attendance = cursor.fetchall()
            
            present_days = 0
            absent_days = 0
            leave_days = 0
            total_working_hours = 0.0
            overtime_hours = 0.0
            
            for record in monthly_attendance:
                if record['status'] == 'present':
                    present_days += 1
                    total_working_hours += record['working_hours'] or daily_hours
                    overtime_hours += record['overtime_hours'] or 0.0
                elif record['status'] == 'absent':
                    absent_days += 1
                elif record['status'] == 'leave':
                    leave_days += 1
            
            # Calculate salary based on monthly salary
            daily_salary = monthly_salary / working_days if working_days > 0 else 0
            basic_salary = present_days * daily_salary
            overtime_pay = (overtime_hours / daily_hours) * daily_salary * 0.5  # 0.5x daily salary for overtime
            deductions = absent_days * daily_salary  # Deduct for absent days
            
            net_salary = basic_salary + overtime_pay - deductions
            
            # Insert or update salary calculation
            cursor.execute('''INSERT OR REPLACE INTO salary_calculations 
                             (userid, month, year, total_days, present_days, absent_days, leave_days,
                              total_working_hours, overtime_hours, basic_salary, overtime_pay, 
                              deductions, net_salary)
                             VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', 
                          (userid, month, year, working_days, present_days, absent_days, leave_days,
                           total_working_hours, overtime_hours, basic_salary, overtime_pay, 
                           deductions, net_salary))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': f'Salary calculated for {month} {year}'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/auto_mark_attendance', methods=['POST'])
@login_required
def api_auto_mark_attendance():
    """Automatically mark attendance based on biometric data"""
    try:
        data = request.get_json()
        date = data.get('date', datetime.now().strftime('%Y-%m-%d'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get all users
        cursor.execute('SELECT userid, shift_start_time, shift_end_time, shift_type, working_hours_per_day FROM users')
        users = cursor.fetchall()
        
        # Check if date is weekend or holiday
        date_obj = datetime.strptime(date, '%Y-%m-%d')
        is_weekend = date_obj.weekday() >= 5
        
        cursor.execute('SELECT * FROM holidays WHERE date = ?', (date,))
        holiday = cursor.fetchone()
        
        marked_count = 0
        
        for user in users:
            userid = user['userid']
            
            # Skip if weekend or holiday (unless user is marked as working)
            if (is_weekend or holiday) and not holiday:
                # Mark as weekend/holiday
                cursor.execute('''INSERT OR REPLACE INTO attendance_marking 
                                 (userid, date, status, working_hours, remarks)
                                 VALUES (?, ?, ?, ?, ?)''', 
                              (userid, date, 'weekend' if is_weekend else 'holiday', 0.0, 
                               'Weekend' if is_weekend else f"Holiday: {holiday['name']}"))
                marked_count += 1
                continue
            
            # Get biometric attendance for this user and date
            cursor.execute('''SELECT check_in, check_out, working_hours 
                             FROM attendance 
                             WHERE userid = ? AND DATE(timestamp) = ?''', (userid, date))
            
            attendance_record = cursor.fetchone()
            
            if attendance_record and attendance_record['check_in'] and attendance_record['check_out']:
                # User has biometric attendance
                working_hours = attendance_record['working_hours'] or 0.0
                daily_hours = user['working_hours_per_day']
                
                # Determine status based on working hours
                if working_hours >= daily_hours * 0.8:  # 80% of required hours
                    status = 'present'
                elif working_hours >= daily_hours * 0.5:  # 50% of required hours
                    status = 'half_day'
                else:
                    status = 'late'
                
                # Calculate overtime
                overtime_hours = max(0, working_hours - daily_hours)
                
                # Calculate late minutes
                late_minutes = 0
                if status == 'late':
                    late_minutes = int((daily_hours - working_hours) * 60)
                
                cursor.execute('''INSERT OR REPLACE INTO attendance_marking 
                                 (userid, date, status, working_hours, overtime_hours, late_minutes, remarks)
                                 VALUES (?, ?, ?, ?, ?, ?, ?)''', 
                              (userid, date, status, working_hours, overtime_hours, late_minutes, 'Auto-marked from biometric'))
                
                marked_count += 1
            else:
                # No biometric attendance - mark as absent
                cursor.execute('''INSERT OR REPLACE INTO attendance_marking 
                                 (userid, date, status, working_hours, remarks)
                                 VALUES (?, ?, ?, ?, ?)''', 
                              (userid, date, 'absent', 0.0, 'No biometric attendance'))
                
                marked_count += 1
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': f'Automatically marked attendance for {marked_count} users'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/monthly_attendance')
@login_required
def api_monthly_attendance():
    """Get monthly attendance data for the monthly sheet view"""
    try:
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        company_id = request.args.get('company_id', type=int)
        employee_id = request.args.get('employee_id', type=int)
        
        if not year or not month:
            return jsonify({'success': False, 'message': 'Missing year or month'})
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Build query for users
        user_query = 'SELECT userid, name, company_name, shift_type, created_date FROM users'
        user_params = []
        
        if company_id and employee_id:
            user_query += ' WHERE company_id = ? AND userid = ?'
            user_params.extend([company_id, employee_id])
        elif company_id:
            user_query += ' WHERE company_id = ?'
            user_params.append(company_id)
        elif employee_id:
            user_query += ' WHERE userid = ?'
            user_params.append(employee_id)
        
        user_query += ' ORDER BY userid'
        
        cursor.execute(user_query, user_params)
        users = cursor.fetchall()
        
        # Get holidays for the month
        month_start = f"{year:04d}-{month:02d}-01"
        if month == 12:
            month_end = f"{year+1:04d}-01-01"
        else:
            month_end = f"{year:04d}-{month+1:02d}-01"
        
        cursor.execute('''SELECT date, name FROM holidays 
                         WHERE date >= ? AND date < ?''', (month_start, month_end))
        holidays = cursor.fetchall()
        
        # Get attendance marking for the month
        attendance_data = []
        for user in users:
            user_attendance = {
                'userid': user['userid'],
                'name': user['name'],
                'company_name': user['company_name'],
                'shift_type': user['shift_type'],
                'created_date': user['created_date'],
                'year': year,
                'month': month,
                'daily_attendance': []
            }
            
            # Get attendance for each day of the month
            days_in_month = (datetime(year, month + 1, 1) - datetime(year, month, 1)).days
            
            for day in range(1, days_in_month + 1):
                date_str = f"{year:04d}-{month:02d}-{day:02d}"
                
                # Check if date is before employee join date
                if user['created_date'] and date_str < user['created_date']:
                    # Employee hasn't joined yet - mark as NA
                    user_attendance['daily_attendance'].append({
                        'date': date_str,
                        'status': 'na',
                        'working_hours': 0.0,
                        'overtime_hours': 0.0,
                        'late_minutes': 0,
                        'remarks': 'Not joined yet',
                        'check_in': None,
                        'check_out': None
                    })
                    continue
                
                # First check if there's manual marking
                cursor.execute('''SELECT status, working_hours, overtime_hours, late_minutes, remarks
                                 FROM attendance_marking 
                                 WHERE userid = ? AND date = ?''', (user['userid'], date_str))
                
                manual_marking = cursor.fetchone()
                
                # Then check actual attendance records
                cursor.execute('''SELECT check_in, check_out, working_hours
                                 FROM attendance 
                                 WHERE userid = ? AND DATE(timestamp) = ?''', (user['userid'], date_str))
                
                actual_attendance = cursor.fetchone()
                
                if manual_marking:
                    # Use manual marking if available
                    print(f"Using manual marking for user {user['userid']} on {date_str}: {manual_marking['status']}")  # Debug log
                    user_attendance['daily_attendance'].append({
                        'date': date_str,
                        'status': manual_marking['status'],
                        'working_hours': manual_marking['working_hours'],
                        'overtime_hours': manual_marking['overtime_hours'],
                        'late_minutes': manual_marking['late_minutes'],
                        'remarks': manual_marking['remarks'],
                        'check_in': None,
                        'check_out': None
                    })
                elif actual_attendance:
                    # Use actual attendance records
                    print(f"Using actual attendance for user {user['userid']} on {date_str}")  # Debug log
                    user_attendance['daily_attendance'].append({
                        'date': date_str,
                        'status': '',  # Will be auto-marked by frontend
                        'working_hours': actual_attendance['working_hours'] or 0.0,
                        'overtime_hours': 0.0,
                        'late_minutes': 0,
                        'remarks': '',
                        'check_in': actual_attendance['check_in'],
                        'check_out': actual_attendance['check_out']
                    })
                else:
                    # No attendance data
                    print(f"No attendance data for user {user['userid']} on {date_str}")  # Debug log
                    user_attendance['daily_attendance'].append({
                        'date': date_str,
                        'status': '',
                        'working_hours': 0.0,
                        'overtime_hours': 0.0,
                        'late_minutes': 0,
                        'remarks': '',
                        'check_in': None,
                        'check_out': None
                    })
            
            attendance_data.append(user_attendance)
        
        conn.close()
        
        # Debug: Print what we're returning
        print(f"Returning attendance data for {len(attendance_data)} users")
        for user_data in attendance_data[:2]:  # Show first 2 users
            print(f"User {user_data['userid']} ({user_data['name']}): {len(user_data['daily_attendance'])} days")
            for day in user_data['daily_attendance'][:5]:  # Show first 5 days
                if day['status']:
                    print(f"  {day['date']}: {day['status']}")
        
        return jsonify({
            'success': True,
            'attendance': attendance_data,
            'holidays': [{'date': h['date'], 'name': h['name']} for h in holidays]
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/export_salary_excel', methods=['POST'])
def export_salary_excel():
    """Export salary data to Excel"""
    try:
        data = request.get_json()
        month = data.get('month')
        year = data.get('year')
        company_id = data.get('company_id')
        
        if not month or not year:
            return jsonify({'success': False, 'message': 'Month and year are required'})
        
        # Get salary data
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Build query based on filters
        query = '''
            SELECT u.userid, u.name, u.company_name, u.monthly_salary,
                   COUNT(DISTINCT DATE(a.timestamp)) as present_days,
                   COUNT(DISTINCT DATE(a.timestamp)) as total_days,
                   SUM(a.working_hours) as total_working_hours,
                   SUM(CASE WHEN a.working_hours > 8 THEN a.working_hours - 8 ELSE 0 END) as overtime_hours
            FROM users u
            LEFT JOIN attendance a ON u.userid = a.userid 
                AND strftime('%Y-%m', a.timestamp) = ?
            WHERE 1=1
        '''
        params = [f"{year}-{month:02d}"]
        
        if company_id:
            query += ' AND u.company_name = (SELECT name FROM companies WHERE id = ?)'
            params.append(company_id)
        
        query += ' GROUP BY u.userid, u.name, u.company_name, u.monthly_salary'
        query += ' ORDER BY u.name'
        
        cursor.execute(query, params)
        records = cursor.fetchall()
        conn.close()
        
        if not records:
            return jsonify({'success': False, 'message': 'No salary data found for the selected period'})
        
        # Generate Excel file
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Salary Report - {month} {year}"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add title
        ws.merge_cells('A1:L1')
        title_cell = ws['A1']
        title_cell.value = f"Salary Report - {month} {year}"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Add company info
        ws.merge_cells('A2:L2')
        company_cell = ws['A2']
        company_cell.value = f"Company: {COMPANY_NAME}"
        company_cell.font = Font(size=12)
        company_cell.alignment = Alignment(horizontal="center")
        
        # Add headers
        headers = [
            'ID', 'Name', 'Company', 'Monthly Salary', 'Present Days', 'Total Days',
            'Working Hours', 'Overtime Hours', 'Basic Salary', 'Overtime Pay', 'Deductions', 'Net Salary'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Add data
        for row, record in enumerate(records, 5):
            # Calculate salary components
            monthly_salary = record['monthly_salary'] or 0
            present_days = record['present_days'] or 0
            total_days = record['total_days'] or 0
            working_hours = record['total_working_hours'] or 0
            overtime_hours = record['overtime_hours'] or 0
            
            # Calculate daily salary
            daily_salary = monthly_salary / 30 if monthly_salary > 0 else 0
            
            # Calculate basic salary
            basic_salary = present_days * daily_salary
            
            # Calculate overtime pay
            overtime_pay = overtime_hours * daily_salary * 0.5
            
            # Calculate deductions
            absent_days = 30 - present_days
            deductions = absent_days * daily_salary
            
            # Calculate net salary
            net_salary = basic_salary + overtime_pay - deductions
            
            # Add row data
            row_data = [
                record['userid'],
                record['name'],
                record['company_name'],
                monthly_salary,
                present_days,
                total_days,
                round(working_hours, 2),
                round(overtime_hours, 2),
                round(basic_salary, 2),
                round(overtime_pay, 2),
                round(deductions, 2),
                round(net_salary, 2)
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Format numbers
                if col in [4, 9, 10, 11, 12]:  # Salary columns
                    cell.number_format = '#,##0.00'
                elif col in [7, 8]:  # Hours columns
                    cell.number_format = '0.00'
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Return Excel file as response
        from flask import send_file
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'salary_report_{month}_{year}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error generating Excel: {str(e)}'})

@app.route('/api/employee_details/<int:user_id>')
@login_required
def get_employee_details(user_id):
    """Get employee details for modal"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT userid, name, company_name, shift_type, monthly_salary, created_date
            FROM users 
            WHERE userid = ?
        ''', (user_id,))
        
        employee = cursor.fetchone()
        conn.close()
        
        if employee:
            return jsonify({
                'success': True,
                'employee': {
                    'userid': employee['userid'],
                    'name': employee['name'],
                    'company_name': employee['company_name'],
                    'shift_type': employee['shift_type'],
                    'monthly_salary': employee['monthly_salary'],
                    'created_date': employee['created_date']
                }
            })
        else:
            return jsonify({'success': False, 'message': 'Employee not found'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/employee_attendance/<int:user_id>/<int:year>/<int:month>')
@login_required
def get_employee_attendance(user_id, year, month):
    """Get employee attendance for a specific month"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get attendance records for the month
        start_date = f"{year:04d}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1:04d}-01-01"
        else:
            end_date = f"{year:04d}-{month + 1:02d}-01"
        
        cursor.execute('''
            SELECT DATE(timestamp) as date, check_in, check_out, working_hours
            FROM attendance 
            WHERE userid = ? AND DATE(timestamp) >= ? AND DATE(timestamp) < ?
            ORDER BY timestamp
        ''', (user_id, start_date, end_date))
        
        records = cursor.fetchall()
        conn.close()
        
        # Process records
        attendance_data = []
        for record in records:
            attendance_data.append({
                'date': record['date'],
                'check_in': record['check_in'],
                'check_out': record['check_out'],
                'working_hours': record['working_hours']
            })
        
        return jsonify({
            'success': True,
            'attendance': attendance_data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/export_employee_attendance/<int:user_id>/<int:year>/<int:month>')
@login_required
def export_employee_attendance(user_id, year, month):
    """Export employee attendance for a specific month"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get employee info
        cursor.execute('SELECT name, company_name FROM users WHERE userid = ?', (user_id,))
        employee = cursor.fetchone()
        
        if not employee:
            return jsonify({'success': False, 'message': 'Employee not found'})
        
        # Get attendance records for the month
        start_date = f"{year:04d}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1:04d}-01-01"
        else:
            end_date = f"{year:04d}-{month + 1:02d}-01"
        
        cursor.execute('''
            SELECT DATE(timestamp) as date, check_in, check_out, working_hours
            FROM attendance 
            WHERE userid = ? AND DATE(timestamp) >= ? AND DATE(timestamp) < ?
            ORDER BY timestamp
        ''', (user_id, start_date, end_date))
        
        records = cursor.fetchall()
        conn.close()
        
        if not records:
            return jsonify({'success': False, 'message': 'No attendance data found for this month'})
        
        # Create Excel workbook
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance - {employee['name']}"
        
        # Add title
        ws['A1'] = f"Monthly Attendance Report - {employee['name']}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Add employee info
        ws['A2'] = f"Employee: {employee['name']}"
        ws['A3'] = f"Company: {employee['company_name']}"
        ws['A4'] = f"Period: {year}-{month:02d}"
        
        # Add headers
        headers = ['Date', 'Check In', 'Check Out', 'Working Hours']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for row, record in enumerate(records, 7):
            ws.cell(row=row, column=1, value=record['date'])
            ws.cell(row=row, column=2, value=record['check_in'])
            ws.cell(row=row, column=3, value=record['check_out'])
            ws.cell(row=row, column=4, value=record['working_hours'] or 0)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temporary file
        import tempfile
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        return send_file(tmp_path, as_attachment=True, 
                        download_name=f'attendance_{employee["name"]}_{year}_{month:02d}.xlsx')
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error exporting attendance: {str(e)}'})

@app.route('/api/employees')
@login_required
def get_employees():
    """Get employees for holiday assignment"""
    try:
        company_id = request.args.get('company_id', type=int)
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if company_id:
            cursor.execute('''
                SELECT userid, name, company_name, shift_type 
                FROM users 
                WHERE company_id = ? 
                ORDER BY name
            ''', (company_id,))
        else:
            cursor.execute('''
                SELECT userid, name, company_name, shift_type 
                FROM users 
                ORDER BY name
            ''')
        
        employees = cursor.fetchall()
        conn.close()
        
        return jsonify({
            'success': True,
            'employees': [{
                'userid': emp['userid'],
                'name': emp['name'],
                'company_name': emp['company_name'],
                'shift_type': emp['shift_type']
            } for emp in employees]
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/assign_holiday', methods=['POST'])
@login_required
def assign_holiday():
    """Assign holiday to specific employees or company"""
    try:
        data = request.get_json()
        date = data.get('date')
        name = data.get('name')
        company_id = data.get('company_id')
        employee_ids = data.get('employee_ids', [])
        
        if not date or not name:
            return jsonify({'success': False, 'message': 'Date and name are required'})
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # If specific employees are selected, assign holiday to them
        if employee_ids:
            for userid in employee_ids:
                # Check if holiday already exists for this employee and date
                cursor.execute('''
                    SELECT id FROM attendance_marking 
                    WHERE userid = ? AND date = ? AND status = 'holiday'
                ''', (userid, date))
                
                existing = cursor.fetchone()
                
                if existing:
                    # Update existing holiday marking
                    cursor.execute('''
                        UPDATE attendance_marking 
                        SET remarks = ? 
                        WHERE userid = ? AND date = ? AND status = 'holiday'
                    ''', (name, userid, date))
                else:
                    # Create new holiday marking
                    cursor.execute('''
                        INSERT INTO attendance_marking 
                        (userid, date, status, working_hours, overtime_hours, late_minutes, remarks)
                        VALUES (?, ?, 'holiday', 0, 0, 0, ?)
                    ''', (userid, date, name))
        
        # If company is selected, assign holiday to all employees in that company
        elif company_id:
            cursor.execute('''
                SELECT userid FROM users WHERE company_id = ?
            ''', (company_id,))
            
            company_employees = cursor.fetchall()
            
            for emp in company_employees:
                userid = emp['userid']
                
                # Check if holiday already exists
                cursor.execute('''
                    SELECT id FROM attendance_marking 
                    WHERE userid = ? AND date = ? AND status = 'holiday'
                ''', (userid, date))
                
                existing = cursor.fetchone()
                
                if not existing:
                    # Create new holiday marking
                    cursor.execute('''
                        INSERT INTO attendance_marking 
                        (userid, date, status, working_hours, overtime_hours, late_minutes, remarks)
                        VALUES (?, ?, 'holiday', 0, 0, 0, ?)
                    ''', (userid, date, name))
        
        # If no specific employees or company, assign to all employees
        else:
            cursor.execute('SELECT userid FROM users')
            all_employees = cursor.fetchall()
            
            for emp in all_employees:
                userid = emp['userid']
                
                # Check if holiday already exists
                cursor.execute('''
                    SELECT id FROM attendance_marking 
                    WHERE userid = ? AND date = ? AND status = 'holiday'
                ''', (userid, date))
                
                existing = cursor.fetchone()
                
                if not existing:
                    # Create new holiday marking
                    cursor.execute('''
                        INSERT INTO attendance_marking 
                        (userid, date, status, working_hours, overtime_hours, late_minutes, remarks)
                        VALUES (?, ?, 'holiday', 0, 0, 0, ?)
                    ''', (userid, date, name))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Holiday "{name}" assigned successfully for {date}'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error assigning holiday: {str(e)}'})

@app.route('/api/update_holiday_name', methods=['POST'])
@login_required
def update_holiday_name():
    """Update holiday name for a specific date"""
    try:
        data = request.get_json()
        date = data.get('date')
        name = data.get('name')
        
        if not date or not name:
            return jsonify({'success': False, 'message': 'Date and name are required'})
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Update holiday names in attendance_marking table
        cursor.execute('''
            UPDATE attendance_marking 
            SET remarks = ? 
            WHERE date = ? AND status = 'holiday'
        ''', (name, date))
        
        # Also update the holidays table if it exists
        cursor.execute('''
            UPDATE holidays 
            SET name = ? 
            WHERE date = ?
        ''', (name, date))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Holiday name updated to "{name}" for {date}'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error updating holiday name: {str(e)}'})

@app.route('/api/clear_attendance_marking', methods=['POST'])
@login_required
def clear_attendance_marking():
    """Clear attendance marking for a specific user and date"""
    try:
        data = request.get_json()
        userid = data.get('userid')
        date = data.get('date')
        
        if not userid or not date:
            return jsonify({'success': False, 'message': 'User ID and date are required'})
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Delete the attendance marking
        cursor.execute('DELETE FROM attendance_marking WHERE userid = ? AND date = ?', (userid, date))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Attendance marking cleared for user {userid} on {date}'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error clearing attendance marking: {str(e)}'})

@app.route('/api/test_insert_attendance')
def test_insert_attendance():
    """Test endpoint to manually insert and retrieve attendance data"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Test data
        test_userid = 1
        test_date = '2025-01-01'
        test_status = 'present'
        
        # Insert test data
        cursor.execute('''INSERT OR REPLACE INTO attendance_marking 
                         (userid, date, status, working_hours, overtime_hours, late_minutes, remarks, marked_by)
                         VALUES (?, ?, ?, ?, ?, ?, ?, ?)''', 
                      (test_userid, test_date, test_status, 8.0, 0.0, 0, 'Test data', 'system'))
        
        conn.commit()
        print(f"Test data inserted: userid={test_userid}, date={test_date}, status={test_status}")
        
        # Retrieve test data
        cursor.execute('SELECT * FROM attendance_marking WHERE userid = ? AND date = ?', (test_userid, test_date))
        retrieved = cursor.fetchone()
        
        conn.close()
        
        return jsonify({
            'success': True,
            'inserted': True,
            'retrieved': dict(retrieved) if retrieved else None,
            'message': 'Test data inserted and retrieved successfully'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/test_attendance_marking')
# @login_required  # Temporarily commented out for debugging
def test_attendance_marking():
    """Test endpoint to verify attendance_marking table"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check table structure
        cursor.execute("PRAGMA table_info(attendance_marking)")
        columns = cursor.fetchall()
        
        # Check if table has data
        cursor.execute("SELECT COUNT(*) FROM attendance_marking")
        count = cursor.fetchone()[0]
        
        # Check sample data
        cursor.execute("SELECT * FROM attendance_marking LIMIT 5")
        sample_data = cursor.fetchall()
        
        conn.close()
        
        return jsonify({
            'success': True,
            'table_structure': [{'name': col[1], 'type': col[2]} for col in columns],
            'total_records': count,
            'sample_data': [dict(row) for row in sample_data] if sample_data else []
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error testing table: {str(e)}'})

@app.route('/api/debug_attendance_marking')
def debug_attendance_marking():
    """Debug endpoint to show raw attendance_marking table contents"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Show all records in attendance_marking table
        cursor.execute("SELECT * FROM attendance_marking ORDER BY userid, date LIMIT 20")
        records = cursor.fetchall()
        
        # Show table structure
        cursor.execute("PRAGMA table_info(attendance_marking)")
        columns = cursor.fetchall()
        
        conn.close()
        
        return jsonify({
            'success': True,
            'table_structure': [{'name': col[1], 'type': col[2]} for col in columns],
            'records': [dict(record) for record in records],
            'total_records': len(records)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/echo', methods=['POST'])
def echo():
    """Simple echo endpoint for testing"""
    try:
        data = request.get_json()
        print(f"Echo received: {data}")
        return jsonify({
            'success': True,
            'message': 'Echo successful',
            'received_data': data
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'Echo error: {str(e)}'})

@app.route('/api/export_attendance', methods=['GET'])
@login_required
def export_attendance_csv():
    """Export attendance data to CSV"""
    try:
        from datetime import datetime, timedelta
        
        # Get filter parameters
        user_id = request.args.get('user_id', '')
        from_date = request.args.get('from_date', '')
        to_date = request.args.get('to_date', '')
        company_id = request.args.get('company_id', '')
        
        # Set default date range if not provided
        if not from_date:
            from_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        if not to_date:
            to_date = datetime.now().strftime('%Y-%m-%d')
        
        conn = get_db_connection()
        
        # Build query based on filters
        query = '''SELECT u.name, u.userid, a.timestamp, a.check_in, a.check_out, 
                   a.working_hours, c.name as company_name
                   FROM attendance a 
                   JOIN users u ON a.userid = u.userid 
                   LEFT JOIN companies c ON u.company_id = c.id
                   WHERE DATE(a.timestamp) >= ? AND DATE(a.timestamp) <= ?'''
        
        params = [from_date, to_date]
        
        if user_id:
            query += ' AND u.userid = ?'
            params.append(user_id)
        
        if company_id:
            query += ' AND u.company_id = ?'
            params.append(company_id)
        
        query += ' ORDER BY a.timestamp DESC'
        
        records = conn.execute(query, params).fetchall()
        conn.close()
        
        if not records:
            return jsonify({'success': False, 'message': 'No data found for the specified period'})
        
        # Create CSV content
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        writer.writerow(['Employee Name', 'Employee ID', 'Date', 'Check In', 'Check Out', 'Working Hours', 'Company'])
        
        # Write data
        for record in records:
            writer.writerow([
                record['name'],
                record['userid'],
                record['timestamp'],
                record['check_in'],
                record['check_out'],
                record['working_hours'],
                record['company_name']
            ])
        
        output.seek(0)
        
        from flask import Response
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=attendance_report_{from_date}_to_{to_date}.csv'}
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error exporting attendance: {str(e)}'})

@app.route('/api/export_attendance_excel', methods=['POST'])
@login_required
def export_attendance_excel():
    """Export attendance data to Excel"""
    try:
        data = request.get_json()
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        company_id = data.get('company_id')
        
        if not start_date or not end_date:
            return jsonify({'success': False, 'message': 'Start date and end date are required'})
        
        # Get attendance data
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Build query based on filters
        query = '''
            SELECT u.userid, u.name, u.company_name, u.shift_type,
                   a.timestamp, a.check_in, a.check_out, a.working_hours
            FROM users u
            LEFT JOIN attendance a ON u.userid = a.userid 
                AND DATE(a.timestamp) BETWEEN ? AND ?
            WHERE 1=1
        '''
        params = [start_date, end_date]
        
        if company_id:
            query += ' AND u.company_name = (SELECT name FROM companies WHERE id = ?)'
            params.append(company_id)
        
        query += ' ORDER BY u.name, a.timestamp'
        
        cursor.execute(query, params)
        records = cursor.fetchall()
        conn.close()
        
        if not records:
            return jsonify({'success': False, 'message': 'No attendance data found for the selected period'})
        
        # Generate Excel file
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance Report - {start_date} to {end_date}"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"Attendance Report - {start_date} to {end_date}"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Add company info
        ws.merge_cells('A2:H2')
        company_cell = ws['A2']
        company_cell.value = f"Company: {COMPANY_NAME}"
        company_cell.font = Font(size=12)
        company_cell.alignment = Alignment(horizontal="center")
        
        # Add headers
        headers = [
            'ID', 'Name', 'Company', 'Shift Type', 'Date', 'Check In', 'Check Out', 'Working Hours'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Add data
        for row, record in enumerate(records, 5):
            # Format date and times
            date_str = record['timestamp'].split(' ')[0] if record['timestamp'] else ''
            check_in = record['check_in'].split(' ')[1] if record['check_in'] else ''
            check_out = record['check_out'].split(' ')[1] if record['check_out'] else ''
            
            row_data = [
                record['userid'],
                record['name'],
                record['company_name'],
                record['shift_type'],
                date_str,
                check_in,
                check_out,
                round(record['working_hours'], 2) if record['working_hours'] else 0
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Format numbers
                if col == 8:  # Working hours column
                    cell.number_format = '0.00'
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Return Excel file as response
        from flask import send_file
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'attendance_report_{start_date}_to_{end_date}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error generating Excel: {str(e)}'})

@app.route('/api/export_users_excel', methods=['POST'])
def export_users_excel():
    """Export users data to Excel"""
    try:
        # Get all users
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''SELECT userid, name, company_name, shift_type, shift_start_time, 
                                shift_end_time, working_hours_per_day, monthly_salary, created_date
                         FROM users ORDER BY name''')
        users = cursor.fetchall()
        conn.close()
        
        if not users:
            return jsonify({'success': False, 'message': 'No users found'})
        
        # Generate Excel file
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Users Report"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add title
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = "Employee Master List"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Add company info
        ws.merge_cells('A2:I2')
        company_cell = ws['A2']
        company_cell.value = f"Company: {COMPANY_NAME}"
        company_cell.font = Font(size=12)
        company_cell.alignment = Alignment(horizontal="center")
        
        # Add headers
        headers = [
            'ID', 'Name', 'Company', 'Shift Type', 'Shift Start', 'Shift End', 
            'Daily Hours', 'Monthly Salary', 'Created Date'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Add data
        for row, user in enumerate(users, 5):
            row_data = [
                user['userid'],
                user['name'],
                user['company_name'],
                user['shift_type'].title() if user['shift_type'] else '',
                user['shift_start_time'],
                user['shift_end_time'],
                user['working_hours_per_day'],
                user['monthly_salary'],
                user['created_date']
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Format numbers
                if col == 8:  # Monthly salary column
                    cell.number_format = '#,##0.00'
                elif col == 7:  # Daily hours column
                    cell.number_format = '0.0'
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Return Excel file as response
        from flask import send_file
        return send_file(
            buffer,
            as_attachment=True,
            download_name='users_export.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error generating Excel: {str(e)}'})

@app.route('/api/send_daily_report', methods=['POST'])
def send_daily_report():
    """Send daily attendance report via email"""
    try:
        data = request.get_json()
        recipients = data.get('recipients', [])
        
        if not recipients:
            return jsonify({'success': False, 'message': 'Recipients list is required'})
        
        # Generate attendance summary
        html_content = generate_attendance_summary_email()
        if not html_content:
            return jsonify({'success': False, 'message': 'No attendance data found for today'})
        
        # Send email
        subject = f"Daily Attendance Summary - {datetime.now().strftime('%Y-%m-%d')} - {COMPANY_NAME}"
        success = send_email_notification(subject, recipients, html_content)
        
        if success:
            return jsonify({'success': True, 'message': 'Daily report sent successfully'})
        else:
            return jsonify({'success': False, 'message': 'Failed to send email'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/send_monthly_report', methods=['POST'])
def send_monthly_report():
    """Send monthly salary report via email"""
    try:
        data = request.get_json()
        month = data.get('month')
        year = data.get('year')
        recipients = data.get('recipients', [])
        
        if not month or not year or not recipients:
            return jsonify({'success': False, 'message': 'Month, year, and recipients are required'})
        
        # Generate salary summary
        html_content = generate_salary_summary_email(month, year)
        if not html_content:
            return jsonify({'success': False, 'message': 'No salary data found for the selected period'})
        
        # Send email
        subject = f"Monthly Salary Summary - {month} {year} - {COMPANY_NAME}"
        success = send_email_notification(subject, recipients, html_content)
        
        if success:
            return jsonify({'success': True, 'message': 'Monthly report sent successfully'})
        else:
            return jsonify({'success': False, 'message': 'Failed to send email'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

# Analytics API Endpoints
@app.route('/api/analytics/companies', methods=['GET'])
@login_required
def get_analytics_companies():
    """Get list of companies for analytics filters"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT DISTINCT company_name FROM users ORDER BY company_name')
        companies = [{'company_name': record['company_name']} for record in cursor.fetchall()]
        
        conn.close()
        return jsonify({'success': True, 'companies': companies})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/analytics/quick-stats', methods=['POST'])
@login_required
def get_analytics_quick_stats_api():
    """Get quick statistics for analytics dashboard"""
    try:
        data = request.get_json()
        company = data.get('company')
        from_date = data.get('fromDate')
        to_date = data.get('toDate')
        shift = data.get('shift')
        
        stats = get_analytics_quick_stats(company, from_date, to_date, shift)
        return jsonify({'success': True, 'stats': stats})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/analytics/attendance-trend', methods=['POST'])
@login_required
def get_analytics_attendance_trend_api():
    """Get attendance trend data for charts"""
    try:
        data = request.get_json()
        company = data.get('company')
        from_date = data.get('fromDate')
        to_date = data.get('toDate')
        shift = data.get('shift')
        
        chart_data = get_attendance_trend_data(company, from_date, to_date, shift)
        return jsonify({'success': True, 'data': chart_data})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/analytics/company-distribution', methods=['GET'])
@login_required
def get_analytics_company_distribution_api():
    """Get company distribution data for pie chart"""
    try:
        chart_data = get_company_distribution_data()
        return jsonify({'success': True, 'data': chart_data})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/analytics/working-hours', methods=['POST'])
@login_required
def get_analytics_working_hours_api():
    """Get working hours data for bar chart"""
    try:
        data = request.get_json()
        company = data.get('company')
        from_date = data.get('fromDate')
        to_date = data.get('toDate')
        shift = data.get('shift')
        
        chart_data = get_working_hours_data(company, from_date, to_date, shift)
        return jsonify({'success': True, 'data': chart_data})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/analytics/overtime-trend', methods=['POST'])
@login_required
def get_analytics_overtime_trend_api():
    """Get overtime trend data for area chart"""
    try:
        data = request.get_json()
        company = data.get('company')
        from_date = data.get('fromDate')
        to_date = data.get('toDate')
        shift = data.get('shift')
        
        chart_data = get_overtime_trend_data(company, from_date, to_date, shift)
        return jsonify({'success': True, 'data': chart_data})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/analytics/performance-table', methods=['POST'])
@login_required
def get_analytics_performance_table_api():
    """Get employee performance data for table"""
    try:
        data = request.get_json()
        company = data.get('company')
        from_date = data.get('fromDate')
        to_date = data.get('toDate')
        shift = data.get('shift')
        
        performance_data = get_performance_table_data(company, from_date, to_date, shift)
        return jsonify({'success': True, 'data': performance_data})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/analytics/export-excel', methods=['POST'])
@login_required
def export_analytics_excel():
    """Export analytics data to Excel"""
    try:
        data = request.get_json()
        company = data.get('company')
        from_date = data.get('fromDate')
        to_date = data.get('toDate')
        shift = data.get('shift')
        
        # Get all analytics data
        quick_stats = get_analytics_quick_stats(company, from_date, to_date, shift)
        attendance_trend = get_attendance_trend_data(company, from_date, to_date, shift)
        performance_data = get_performance_table_data(company, from_date, to_date, shift)
        
        # Create Excel workbook
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        
        # Summary sheet
        ws1 = wb.active
        ws1.title = "Summary"
        
        # Add title
        ws1['A1'] = f"Analytics Report - {COMPANY_NAME}"
        ws1['A1'].font = Font(size=16, bold=True)
        ws1.merge_cells('A1:D1')
        
        # Add filters
        ws1['A3'] = "Filters Applied:"
        ws1['A3'].font = Font(bold=True)
        ws1['A4'] = f"Company: {company if company else 'All'}"
        ws1['A5'] = f"Date Range: {from_date} to {to_date}"
        ws1['A6'] = f"Shift: {shift if shift else 'All'}"
        
        # Add quick stats
        ws1['A8'] = "Quick Statistics"
        ws1['A8'].font = Font(bold=True, size=14)
        ws1['A9'] = "Total Employees"
        ws1['B9'] = quick_stats['totalEmployees']
        ws1['A10'] = "Present Today"
        ws1['B10'] = quick_stats['presentToday']
        ws1['A11'] = "Late Today"
        ws1['B11'] = quick_stats['lateToday']
        ws1['A12'] = "Average Working Hours"
        ws1['B12'] = quick_stats['avgWorkingHours']
        
        # Performance sheet
        ws2 = wb.create_sheet("Performance")
        
        # Add headers
        headers = ['Employee', 'Company', 'Attendance Rate %', 'Avg Working Hours', 'Overtime Hours', 'Late Arrivals', 'Performance Score']
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for row, employee in enumerate(performance_data, 2):
            ws2.cell(row=row, column=1, value=employee['name'])
            ws2.cell(row=row, column=2, value=employee['company_name'])
            ws2.cell(row=row, column=3, value=employee['attendanceRate'])
            ws2.cell(row=row, column=4, value=employee['avgWorkingHours'])
            ws2.cell(row=row, column=5, value=employee['overtimeHours'])
            ws2.cell(row=row, column=6, value=employee['lateArrivals'])
            ws2.cell(row=row, column=7, value=employee['performanceScore'])
        
        # Auto-adjust column widths
        for sheet in [ws1, ws2]:
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temporary file
        import tempfile
        import os
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        return send_file(tmp_path, as_attachment=True, 
                        download_name=f'analytics_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error exporting analytics: {str(e)}'})

@app.route('/api/save_email_config', methods=['POST'])
def save_email_config_api():
    """Save email configuration"""
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        
        if not email or not password:
            return jsonify({'success': False, 'message': 'Email and password are required'})
        
        # Save to database
        if save_email_config(email, password):
            return jsonify({'success': True, 'message': 'Email configuration saved successfully'})
        else:
            return jsonify({'success': False, 'message': 'Failed to save email configuration'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/get_email_config', methods=['GET'])
def get_email_config_api():
    """Get saved email configuration"""
    try:
        config = load_email_config()
        if config:
            return jsonify({'success': True, 'config': config})
        else:
            return jsonify({'success': False, 'message': 'No email configuration found'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/send_test_email', methods=['POST'])
def send_test_email():
    """Send test email to verify configuration"""
    try:
        data = request.get_json()
        recipient = data.get('email')
        
        if not recipient:
            return jsonify({'success': False, 'message': 'Email address is required'})
        
        # Get email configuration from request
        email_config = data.get('email_config', {})
        
        # Update app config with the provided email settings
        if email_config.get('email') and email_config.get('password'):
            app.config['MAIL_USERNAME'] = email_config['email']
            app.config['MAIL_PASSWORD'] = email_config['password']
            app.config['MAIL_DEFAULT_SENDER'] = email_config['email']
            
            # Reinitialize Flask-Mail with new config
            global mail
            mail = Mail(app)
            
            # Save configuration to database
            save_email_config(email_config['email'], email_config['password'])
        
        # Simple test email
        html_content = f"""
        <html>
        <body>
            <h2>Test Email from Attendance System</h2>
            <p>This is a test email to verify your email configuration.</p>
            <p>Company: {COMPANY_NAME}</p>
            <p>Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <p>If you receive this email, your email configuration is working correctly!</p>
        </body>
        </html>
        """
        
        subject = f"Test Email - {COMPANY_NAME} Attendance System"
        success = send_email_notification(subject, [recipient], html_content)
        
        if success:
            return jsonify({'success': True, 'message': 'Test email sent successfully'})
        else:
            return jsonify({'success': False, 'message': 'Failed to send test email'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

if __name__ == '__main__':
    # Initialize database
    setup_db()
    
    # Setup automated email schedules
    try:
        # Daily attendance report at 6:00 PM
        scheduler.add_job(
            func=lambda: send_daily_report_automated(),
            trigger=CronTrigger(hour=18, minute=0),
            id='daily_attendance_report',
            name='Send daily attendance report',
            replace_existing=True
        )
        
        # Monthly salary report on 1st of each month at 9:00 AM
        scheduler.add_job(
            func=lambda: send_monthly_report_automated(),
            trigger=CronTrigger(day=1, hour=9, minute=0),
            id='monthly_salary_report',
            name='Send monthly salary report',
            replace_existing=True
        )
        
        print("✅ Automated email schedules configured successfully!")
        
    except Exception as e:
        print(f"⚠️ Warning: Could not setup automated emails: {e}")
        print("You can still send manual reports via the web interface")
    
    print("Starting Attendance System Web App...")
    print("Open your browser and go to: http://localhost:5000")
    print("Press Ctrl+C to stop the server")
    
    app.run(host='0.0.0.0', port=5000, debug=True)

def send_daily_report_automated():
    """Automated daily attendance report (called by scheduler)"""
    try:
        # Default recipients - you can change these
        default_recipients = ['manager@company.com', 'hr@company.com']
        
        html_content = generate_attendance_summary_email()
        if html_content:
            subject = f"Daily Attendance Summary - {datetime.now().strftime('%Y-%m-%d')} - {COMPANY_NAME}"
            send_email_notification(subject, default_recipients, html_content)
            print(f"✅ Automated daily report sent at {datetime.now()}")
        else:
            print(f"⚠️ No attendance data for automated daily report at {datetime.now()}")
            
    except Exception as e:
        print(f"❌ Error sending automated daily report: {e}")

def send_monthly_report_automated():
    """Automated monthly salary report (called by scheduler)"""
    try:
        # Default recipients - you can change these
        default_recipients = ['hr@company.com', 'payroll@company.com']
        
        # Get current month and year
        now = datetime.now()
        month = now.month
        year = now.year
        
        html_content = generate_salary_summary_email(month, year)
        if html_content:
            subject = f"Monthly Salary Summary - {month} {year} - {COMPANY_NAME}"
            send_email_notification(subject, default_recipients, html_content)
            print(f"✅ Automated monthly report sent at {datetime.now()}")
        else:
            print(f"⚠️ No salary data for automated monthly report at {datetime.now()}")
            
    except Exception as e:
        print(f"❌ Error sending automated monthly report: {e}")
